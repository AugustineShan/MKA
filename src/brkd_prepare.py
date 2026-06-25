"""Prepare BRKD source materials as Markdown.

The /brkd AI step should read only deterministic Markdown, not raw PDF/Word/Excel
files.  This module scans a company's
``Skills素材包/BRKD业务理解器（研报和纪要放在这里）`` directory, converts supported
top-level files into ``markdown存储区/``, and writes a manifest for audit.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import zipfile
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import fitz  # PyMuPDF
from openpyxl import load_workbook

from src.company_paths import (
    COMPANIES_DIR,
    brkd_markdown_store_dir,
    brkd_material_dir,
    find_company_dir,
)


CONVERTER_VERSION = 1
SUPPORTED_SUFFIXES = {
    ".pdf",
    ".md",
    ".markdown",
    ".txt",
    ".csv",
    ".tsv",
    ".docx",
    ".doc",
    ".xlsx",
    ".xlsm",
    ".xls",
}
TEXT_SUFFIXES = {".md", ".markdown", ".txt"}
TABLE_TEXT_SUFFIXES = {".csv", ".tsv"}
EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
UNSUPPORTED_SUFFIXES = {".doc", ".xls"}
MAX_EXCEL_ROWS = 800
MAX_EXCEL_COLS = 80


class BrkdPrepareError(RuntimeError):
    """Raised when BRKD materials cannot be prepared."""


@dataclass(frozen=True)
class MaterialResult:
    source_path: str
    markdown_path: str
    status: str
    converter: str
    message: str = ""


def resolve_company(raw: str | Path, *, companies_dir: Path = COMPANIES_DIR) -> Path:
    path = Path(raw)
    if path.exists() and path.is_dir():
        return path

    text = str(raw).strip()
    if not text:
        raise BrkdPrepareError("company argument is empty")

    if re.fullmatch(r"\d{6}(?:\.(?:SZ|SH|BJ))?", text.upper()):
        return find_company_dir(text, companies_dir)

    candidates = sorted(companies_dir.glob(f"{text}_*"))
    if len(candidates) == 1:
        return candidates[0]
    if len(candidates) > 1:
        names = ", ".join(path.name for path in candidates)
        raise BrkdPrepareError(f"multiple company directories match {text}: {names}")
    raise BrkdPrepareError(f"no company directory matches {text}")


def _source_files(source_dir: Path) -> list[Path]:
    if not source_dir.exists():
        return []
    files: list[Path] = []
    for path in sorted(source_dir.iterdir(), key=lambda p: p.name.lower()):
        if not path.is_file():
            continue
        if path.name.startswith("~$"):
            continue
        if path.suffix.lower() not in SUPPORTED_SUFFIXES:
            continue
        files.append(path)
    return files


def _safe_output_name(path: Path) -> str:
    stem = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", path.stem).strip(" ._") or "material"
    ext = path.suffix.lower().lstrip(".") or "file"
    return f"{stem}__{ext}.md"


def _frontmatter_value(value: str) -> str:
    return json.dumps(value, ensure_ascii=False)


def _frontmatter(source: Path, *, converter: str, status: str, message: str = "") -> str:
    stat = source.stat()
    lines = [
        "---",
        f"source_file: {_frontmatter_value(source.name)}",
        f"source_path: {_frontmatter_value(str(source))}",
        f"source_suffix: {_frontmatter_value(source.suffix.lower())}",
        f"source_size: {stat.st_size}",
        f"source_mtime_ns: {stat.st_mtime_ns}",
        f"converted_at: {_frontmatter_value(dt.datetime.now().isoformat(timespec='seconds'))}",
        f"converter: {_frontmatter_value(converter)}",
        f"converter_version: {CONVERTER_VERSION}",
        f"status: {_frontmatter_value(status)}",
    ]
    if message:
        lines.append(f"message: {_frontmatter_value(message)}")
    lines.extend(["---", ""])
    return "\n".join(lines)


def _read_existing_signature(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    text = path.read_text(encoding="utf-8", errors="replace")
    if not text.startswith("---\n"):
        return {}
    end = text.find("\n---\n", 4)
    if end == -1:
        return {}
    result: dict[str, str] = {}
    for line in text[4:end].splitlines():
        if ":" not in line:
            continue
        key, value = line.split(":", 1)
        result[key.strip()] = value.strip().strip('"')
    return result


def _is_current(source: Path, md_path: Path) -> bool:
    sig = _read_existing_signature(md_path)
    if not sig:
        return False
    stat = source.stat()
    return (
        sig.get("source_size") == str(stat.st_size)
        and sig.get("source_mtime_ns") == str(stat.st_mtime_ns)
        and sig.get("converter_version") == str(CONVERTER_VERSION)
    )


def _read_text(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "utf-16"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _convert_pdf(path: Path) -> tuple[str, str]:
    parts: list[str] = []
    with fitz.open(path) as doc:
        for index, page in enumerate(doc, start=1):
            parts.append(f"## 第 {index} 页")
            parts.append("")
            parts.append(page.get_text("text").strip())
            parts.append("")
    return "\n".join(parts).strip() + "\n", "PyMuPDF(fitz) plain-text"


def _convert_text(path: Path) -> tuple[str, str]:
    body = _read_text(path).strip()
    if path.suffix.lower() in {".md", ".markdown"}:
        return body + "\n", "markdown-copy"
    return f"```text\n{body}\n```\n", "text-copy"


def _convert_delimited_text(path: Path) -> tuple[str, str]:
    body = _read_text(path).strip()
    fence = "tsv" if path.suffix.lower() == ".tsv" else "csv"
    return f"```{fence}\n{body}\n```\n", f"{fence}-copy"


def _convert_docx(path: Path) -> tuple[str, str]:
    try:
        with zipfile.ZipFile(path) as archive:
            xml = archive.read("word/document.xml")
    except KeyError as exc:
        raise BrkdPrepareError(f"docx missing word/document.xml: {path}") from exc

    root = ET.fromstring(xml)
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs: list[str] = []
    for para in root.findall(".//w:p", namespace):
        texts = [node.text or "" for node in para.findall(".//w:t", namespace)]
        text = "".join(texts).strip()
        if text:
            paragraphs.append(text)
    return "\n\n".join(paragraphs).strip() + "\n", "docx-xml"


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    return text.replace("|", "\\|")


def _convert_excel(path: Path) -> tuple[str, str]:
    wb = load_workbook(path, data_only=True, read_only=True)
    parts: list[str] = []
    for ws in wb.worksheets:
        parts.append(f"## Sheet: {ws.title}")
        parts.append("")
        rows_written = 0
        truncated = False
        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_index > MAX_EXCEL_ROWS:
                truncated = True
                break
            values = [_cell_text(cell) for cell in row[:MAX_EXCEL_COLS]]
            while values and values[-1] == "":
                values.pop()
            if not values:
                continue
            rows_written += 1
            parts.append("| " + " | ".join(values) + " |")
        if rows_written == 0:
            parts.append("_空 sheet_")
        if truncated:
            parts.append("")
            parts.append(f"> 已截断：仅保留前 {MAX_EXCEL_ROWS} 行。")
        parts.append("")
    return "\n".join(parts).strip() + "\n", "openpyxl(data_only=True)"


def _unsupported(path: Path) -> tuple[str, str, str]:
    suffix = path.suffix.lower()
    if suffix == ".doc":
        message = "旧版 .doc 暂不支持确定性转换；请另存为 .docx 或 PDF 后重新运行 brkd prepare。"
    elif suffix == ".xls":
        message = "旧版 .xls 暂不支持确定性转换；请另存为 .xlsx 或 .xlsm 后重新运行 brkd prepare。"
    else:
        message = f"暂不支持该文件类型: {suffix}"
    body = f"# 转换未完成\n\n{message}\n"
    return body, "unsupported-placeholder", message


def _convert_body(path: Path) -> tuple[str, str, str, str]:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        body, converter = _convert_pdf(path)
        return body, converter, "converted", ""
    if suffix in TEXT_SUFFIXES:
        body, converter = _convert_text(path)
        return body, converter, "converted", ""
    if suffix in TABLE_TEXT_SUFFIXES:
        body, converter = _convert_delimited_text(path)
        return body, converter, "converted", ""
    if suffix == ".docx":
        body, converter = _convert_docx(path)
        return body, converter, "converted", ""
    if suffix in EXCEL_SUFFIXES:
        body, converter = _convert_excel(path)
        return body, converter, "converted", ""
    if suffix in UNSUPPORTED_SUFFIXES:
        body, converter, message = _unsupported(path)
        return body, converter, "unsupported", message
    body, converter, message = _unsupported(path)
    return body, converter, "unsupported", message


def convert_material(
    source: Path,
    markdown_dir: Path,
    *,
    force: bool = False,
    output_name: str | None = None,
) -> MaterialResult:
    markdown_dir.mkdir(parents=True, exist_ok=True)
    md_path = markdown_dir / (output_name or _safe_output_name(source))
    if md_path.exists() and not force and _is_current(source, md_path):
        return MaterialResult(str(source), str(md_path), "skipped", "unchanged")

    try:
        body, converter, status, message = _convert_body(source)
    except Exception as exc:  # noqa: BLE001 - preserve error as auditable markdown.
        converter = "conversion-error"
        status = "error"
        message = f"{type(exc).__name__}: {exc}"
        body = f"# 转换失败\n\n{message}\n"

    title = f"# {source.name}\n\n"
    md_path.write_text(
        _frontmatter(source, converter=converter, status=status, message=message) + title + body,
        encoding="utf-8",
    )
    return MaterialResult(str(source), str(md_path), status, converter, message)


def prepare_brkd_materials(
    company: str | Path | None = None,
    *,
    folder: str | Path | None = None,
    force: bool = False,
    companies_dir: Path = COMPANIES_DIR,
) -> dict[str, Any]:
    if folder is not None:
        source_dir = Path(folder)
        company_dir = source_dir.parent.parent if source_dir.name.startswith("BRKD") else None
        markdown_dir = source_dir / "markdown存储区"
    elif company is not None:
        company_dir = resolve_company(company, companies_dir=companies_dir)
        source_dir = brkd_material_dir(company_dir)
        markdown_dir = brkd_markdown_store_dir(company_dir)
    else:
        raise BrkdPrepareError("provide company or folder")

    if not source_dir.exists():
        raise BrkdPrepareError(f"BRKD material directory does not exist: {source_dir}")

    sources = _source_files(source_dir)
    markdown_dir.mkdir(parents=True, exist_ok=True)
    results = [convert_material(source, markdown_dir, force=force) for source in sources]
    counts: dict[str, int] = {}
    for result in results:
        counts[result.status] = counts.get(result.status, 0) + 1

    manifest = {
        "mode": "brkd_prepare",
        "company_dir": str(company_dir) if company_dir is not None else None,
        "source_dir": str(source_dir),
        "markdown_dir": str(markdown_dir),
        "force": force,
        "source_count": len(sources),
        "counts": counts,
        "materials": [asdict(result) for result in results],
    }
    (markdown_dir / "brkd_prepare_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return manifest


def _print_manifest(manifest: dict[str, Any]) -> None:
    print(f"BRKD source dir: {manifest['source_dir']}")
    print(f"Markdown store: {manifest['markdown_dir']}")
    print(f"Sources: {manifest['source_count']}")
    for status, count in sorted((manifest.get("counts") or {}).items()):
        print(f"{status}: {count}")
    for item in manifest.get("materials", []):
        print(f"- [{item['status']}] {Path(item['source_path']).name} -> {Path(item['markdown_path']).name}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Prepare BRKD materials into markdown存储区")
    parser.add_argument("company", nargs="?", help="公司名、裸代码、完整 ticker 或公司目录")
    parser.add_argument("--folder", type=Path, help="直接指定 BRKD 素材文件夹")
    parser.add_argument("--force", action="store_true", help="强制重新转换")
    args = parser.parse_args()

    try:
        manifest = prepare_brkd_materials(args.company, folder=args.folder, force=args.force)
    except Exception as exc:  # noqa: BLE001 - CLI should report concise failure.
        print(f"BRKD prepare failed: {exc}")
        return 2
    _print_manifest(manifest)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
