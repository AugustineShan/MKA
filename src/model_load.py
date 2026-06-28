"""Vintage model loading helpers for the /load skill.

This module deliberately keeps /load outputs isolated from the official
``Agent/forecast`` pipeline.  It builds a time-capped sandbox from an external
workbook model, then runs DCF only inside ``Agent/Load/<load_id>/forecast``.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.company_paths import (
    COMPANIES_DIR,
    agent_dir,
    annual_reports_dir,
    db_path as company_db_path,
    find_company_dir,
    ka_reference_dir,
    load_model_dir,
)
from src.defaults_gen import build_defaults
from src.forecast import ForecastRun, run_company_forecast
from src.yaml2_schema import write_yaml2


EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls"}
YEAR_RE = re.compile(r"(?:19|20)\d{2}")
DATE8_RE = re.compile(r"(?<!\d)((?:19|20)\d{6})(?!\d)")
DATE6_RE = re.compile(r"(?<!\d)(\d{6})(?!\d)")
FORECAST_LABEL_RE = re.compile(r"^\s*((?:19|20)\d{2})\s*(?:E|F|A/E|E/A|预测|預測)", re.IGNORECASE)
HEADER_YEAR_MIN_COUNT = 3


class ModelLoadError(RuntimeError):
    """Raised when /load cannot safely establish a vintage sandbox."""


@dataclass(frozen=True)
class YearLabel:
    column: int
    year: int
    raw: str
    is_forecast: bool

    def as_dict(self) -> dict[str, Any]:
        return {
            "column": self.column,
            "year": self.year,
            "raw": self.raw,
            "is_forecast": self.is_forecast,
        }


@dataclass(frozen=True)
class HeaderCandidate:
    sheet: str
    row: int
    labels: tuple[YearLabel, ...]
    explicit_forecast_start: int | None = None
    formula_switch_start: int | None = None
    formula_switch_detail: str | None = None

    @property
    def years(self) -> list[int]:
        return [label.year for label in self.labels]

    @property
    def first_forecast_start(self) -> int | None:
        return self.explicit_forecast_start or self.formula_switch_start

    def as_dict(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet,
            "row": self.row,
            "labels": [label.as_dict() for label in self.labels],
            "explicit_forecast_start": self.explicit_forecast_start,
            "formula_switch_start": self.formula_switch_start,
            "formula_switch_detail": self.formula_switch_detail,
        }


@dataclass(frozen=True)
class ModelBoundary:
    model_path: Path
    model_asof_date: str | None
    history_end_year: int
    forecast_start_year: int
    # audit C1·同名异义警示:这里 forecast_years 是【预测年份列表】(如 (2025,2026,...))。
    # 注意 calc.py / yaml2_schema.py / da_roll.py / forecast.py 里的 `forecast_years` 是【年数(int)】,
    # 两者同名不同义、分属 LOAD 层与 DCF 层、不在同一命名空间不会运行期相撞,但读码时勿混。
    forecast_years: tuple[int, ...]
    source: HeaderCandidate
    candidates: tuple[HeaderCandidate, ...]
    conflicts: tuple[str, ...] = ()

    def as_dict(self) -> dict[str, Any]:
        return {
            "model_path": str(self.model_path),
            "model_asof_date": self.model_asof_date,
            "history_end_year": self.history_end_year,
            "forecast_start_year": self.forecast_start_year,
            "forecast_years": list(self.forecast_years),
            "source": self.source.as_dict(),
            "candidates": [candidate.as_dict() for candidate in self.candidates],
            "conflicts": list(self.conflicts),
        }


def resolve_company(raw: str | Path, *, companies_dir: Path = COMPANIES_DIR) -> Path:
    path = Path(raw)
    if path.exists() and path.is_dir():
        return path

    text = str(raw).strip()
    if not text:
        raise ModelLoadError("company argument is empty")

    if re.fullmatch(r"\d{6}(?:\.(?:SZ|SH|BJ))?", text.upper()):
        return find_company_dir(text, companies_dir)

    candidates = sorted(companies_dir.glob(f"{text}_*"))
    if len(candidates) == 1:
        return candidates[0]
    if len(candidates) > 1:
        names = ", ".join(path.name for path in candidates)
        raise ModelLoadError(f"multiple company directories match {text}: {names}")
    raise ModelLoadError(f"no company directory matches {text}")


def _is_model_file(path: Path) -> bool:
    return path.is_file() and not path.name.startswith("~$") and path.suffix.lower() in EXCEL_SUFFIXES


def load_source_model(company_dir: Path) -> Path:
    directory = load_model_dir(company_dir)
    if not directory.exists():
        raise ModelLoadError(f"missing /load model material directory: {directory}")

    candidates = sorted(path for path in directory.iterdir() if _is_model_file(path))
    if not candidates:
        raise ModelLoadError(f"no Excel model found under {directory}")
    if len(candidates) > 1:
        names = ", ".join(path.name for path in candidates)
        raise ModelLoadError(
            f"/load expects exactly one Excel model under {directory}, found {len(candidates)}: {names}"
        )
    return candidates[0]


def latest_active_model(company_dir: Path) -> Path:
    """Backward-compatible alias for older callers; /load now uses the load material folder."""
    return load_source_model(company_dir)


def _year_label(value: Any, column: int) -> YearLabel | None:
    if value is None:
        return None
    if isinstance(value, int) and 1900 <= value <= 2100:
        return YearLabel(column=column, year=value, raw=str(value), is_forecast=False)
    if isinstance(value, float) and abs(value - round(value)) < 1e-9 and 1900 <= int(value) <= 2100:
        year = int(value)
        return YearLabel(column=column, year=year, raw=str(value), is_forecast=False)
    text = str(value).strip()
    forecast_match = FORECAST_LABEL_RE.search(text)
    if forecast_match:
        year = int(forecast_match.group(1))
        return YearLabel(column=column, year=year, raw=text, is_forecast=True)
    match = YEAR_RE.search(text)
    if not match:
        return None
    year = int(match.group(0))
    return YearLabel(column=column, year=year, raw=text, is_forecast=False)


def _sheet_priority(name: str) -> int:
    if "年度和半年度" in name:
        return 0
    if "核心假设" in name:
        return 1
    if name.lower() == "summary":
        return 2
    return 3


def _looks_like_raw_formula(value: Any) -> bool:
    if not isinstance(value, str) or not value.startswith("="):
        return False
    upper = value.upper()
    return "!" in value or "RAW" in upper or "INDEX(" in upper or "VLOOKUP(" in upper or "XLOOKUP(" in upper


def _formula_switch_start(
    ws: Worksheet,
    *,
    header_row: int,
    labels: tuple[YearLabel, ...],
) -> tuple[int | None, str | None]:
    stats: dict[int, dict[str, int]] = {}
    max_row = min(ws.max_row, header_row + 120)
    for label in labels:
        raw_formula_count = 0
        literal_count = 0
        formula_count = 0
        for row in range(header_row + 1, max_row + 1):
            value = ws.cell(row=row, column=label.column).value
            if value is None:
                continue
            if isinstance(value, str) and value.startswith("="):
                formula_count += 1
                if _looks_like_raw_formula(value):
                    raw_formula_count += 1
            else:
                literal_count += 1
        stats[label.column] = {
            "raw_formula": raw_formula_count,
            "formula": formula_count,
            "literal": literal_count,
        }

    for prev_label, label in zip(labels, labels[1:]):
        prev = stats[prev_label.column]
        cur = stats[label.column]
        if prev["raw_formula"] >= 1 and cur["raw_formula"] == 0 and cur["literal"] >= 1:
            detail = (
                f"{ws.title}!R{header_row}: {prev_label.year} has raw/formula refs "
                f"({prev['raw_formula']}), {label.year} switches to literals "
                f"({cur['literal']})"
            )
            return label.year, detail
    return None, None


def _candidate_rows(ws: Worksheet) -> list[HeaderCandidate]:
    candidates: list[HeaderCandidate] = []
    for row in range(1, ws.max_row + 1):
        labels: list[YearLabel] = []
        for column in range(1, ws.max_column + 1):
            label = _year_label(ws.cell(row=row, column=column).value, column)
            if label is not None:
                labels.append(label)
        unique_years = {label.year for label in labels}
        if len(unique_years) < HEADER_YEAR_MIN_COUNT:
            continue
        labels = sorted(labels, key=lambda item: item.column)
        explicit = next((label.year for label in labels if label.is_forecast), None)
        formula_start, formula_detail = _formula_switch_start(
            ws,
            header_row=row,
            labels=tuple(labels),
        )
        candidates.append(
            HeaderCandidate(
                sheet=ws.title,
                row=row,
                labels=tuple(labels),
                explicit_forecast_start=explicit,
                formula_switch_start=formula_start,
                formula_switch_detail=formula_detail,
            )
        )
    return candidates


def _candidate_sort_key(candidate: HeaderCandidate) -> tuple[int, int, int, int]:
    has_forecast = 0 if candidate.first_forecast_start is not None else 1
    return (
        has_forecast,
        _sheet_priority(candidate.sheet),
        -len(candidate.labels),
        candidate.row,
    )


def _asof_from_filename(path: Path) -> str | None:
    text = path.stem
    match = DATE8_RE.search(text)
    if match:
        raw = match.group(1)
        return f"{raw[:4]}-{raw[4:6]}-{raw[6:8]}"
    for match in DATE6_RE.finditer(text):
        raw = match.group(1)
        yy = int(raw[:2])
        year = 2000 + yy if yy < 80 else 1900 + yy
        month = int(raw[2:4])
        day = int(raw[4:6])
        if 1 <= month <= 12 and 1 <= day <= 31:
            return f"{year:04d}-{month:02d}-{day:02d}"
    return None


def inspect_workbook_boundary(model_path: str | Path) -> ModelBoundary:
    path = Path(model_path)
    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    candidates: list[HeaderCandidate] = []
    for sheet_name in workbook.sheetnames:
        candidates.extend(_candidate_rows(workbook[sheet_name]))
    candidates = sorted(candidates, key=_candidate_sort_key)
    source = next((candidate for candidate in candidates if candidate.first_forecast_start is not None), None)
    if source is None:
        raise ModelLoadError(f"cannot infer model forecast start from workbook: {path}")

    conflicts: list[str] = []
    if (
        source.explicit_forecast_start is not None
        and source.formula_switch_start is not None
        and source.explicit_forecast_start != source.formula_switch_start
    ):
        conflicts.append(
            "explicit forecast label starts at "
            f"{source.explicit_forecast_start}, but formula/literal switch starts at "
            f"{source.formula_switch_start} on {source.sheet}!R{source.row}"
        )

    forecast_start = source.first_forecast_start
    if forecast_start is None:
        raise ModelLoadError(f"cannot infer model forecast start from workbook: {path}")
    forecast_years = tuple(year for year in source.years if year >= forecast_start)
    if not forecast_years:
        forecast_years = tuple(range(forecast_start, forecast_start + 5))

    return ModelBoundary(
        model_path=path,
        model_asof_date=_asof_from_filename(path),
        history_end_year=forecast_start - 1,
        forecast_start_year=forecast_start,
        forecast_years=forecast_years,
        source=source,
        candidates=tuple(candidates[:10]),
        conflicts=tuple(conflicts),
    )


def _quote_ident(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def _period_tables(conn: sqlite3.Connection) -> list[str]:
    tables = [
        row[0]
        for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name").fetchall()
    ]
    out: list[str] = []
    for table in tables:
        columns = [row[1] for row in conn.execute(f"PRAGMA table_info({_quote_ident(table)})").fetchall()]
        if "period" in columns:
            out.append(table)
    return out


def _count_after_cutoff(conn: sqlite3.Connection, table: str, history_end_year: int) -> int:
    quoted = _quote_ident(table)
    row = conn.execute(
        f"SELECT COUNT(*) FROM {quoted} WHERE CAST(SUBSTR(CAST(period AS TEXT), 1, 4) AS INTEGER) > ?",
        (history_end_year,),
    ).fetchone()
    return int(row[0])


def copy_cutoff_database(source_db: Path, target_db: Path, history_end_year: int) -> dict[str, int]:
    target_db.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_db, target_db)
    removed: dict[str, int] = {}
    with sqlite3.connect(target_db) as conn:
        for table in _period_tables(conn):
            count = _count_after_cutoff(conn, table, history_end_year)
            if count:
                quoted = _quote_ident(table)
                conn.execute(
                    f"DELETE FROM {quoted} "
                    "WHERE CAST(SUBSTR(CAST(period AS TEXT), 1, 4) AS INTEGER) > ?",
                    (history_end_year,),
                )
                removed[table] = count
        conn.commit()
    return removed


def _year_from_name(path: Path) -> int | None:
    match = YEAR_RE.search(path.name)
    return int(match.group(0)) if match else None


def _annual_materials(company_dir: Path, *, history_end_year: int, forecast_start_year: int) -> tuple[list[Path], list[Path]]:
    allowed: list[Path] = []
    forbidden: list[Path] = []
    reports_dir = annual_reports_dir(company_dir)
    if not reports_dir.exists():
        return allowed, forbidden
    for path in sorted(reports_dir.rglob("*")):
        if not path.is_file():
            continue
        year = _year_from_name(path)
        if year is None:
            continue
        if year <= history_end_year and path.suffix.lower() == ".md":
            allowed.append(path)
        elif year >= forecast_start_year:
            forbidden.append(path)
    return allowed, forbidden


def _copy_allowed_materials(load_dir: Path, model_path: Path, allowed_reports: list[Path]) -> list[Path]:
    target = load_dir / "allowed_materials"
    target.mkdir(parents=True, exist_ok=True)
    copied: list[Path] = []

    model_dest = target / model_path.name
    shutil.copy2(model_path, model_dest)
    copied.append(model_dest)

    reports_dest = target / "annual_reports"
    for report in allowed_reports:
        reports_dest.mkdir(parents=True, exist_ok=True)
        dest = reports_dest / report.name
        shutil.copy2(report, dest)
        copied.append(dest)
    return copied


def _write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.rstrip() + "\n", encoding="utf-8")


def _boundary_markdown(boundary: ModelBoundary) -> str:
    horizon = ", ".join(str(year) for year in boundary.forecast_years)
    return f"""# /load model boundary

- model_path: {boundary.model_path}
- model_asof_date: {boundary.model_asof_date or "unknown"}
- history_end_year: {boundary.history_end_year}
- forecast_start_year: {boundary.forecast_start_year}
- forecast_years: [{horizon}]
- source: {boundary.source.sheet}!R{boundary.source.row}

This file is the load-mode time fence.  Later actuals are future information for
this vintage model.
"""


def _forbidden_markdown(
    boundary: ModelBoundary,
    forbidden_reports: list[Path],
    removed_rows: dict[str, int],
    company_dir: Path,
) -> str:
    lines = [
        "# /load forbidden materials",
        "",
        f"- history_end_year: {boundary.history_end_year}",
        f"- forecast_start_year: {boundary.forecast_start_year}",
        "",
        "## Annual or quarterly reports not allowed",
    ]
    if forbidden_reports:
        lines.extend(f"- {path}" for path in forbidden_reports)
    else:
        lines.append("- none found")
    lines.extend(["", "## Database rows excluded from sandbox"])
    if removed_rows:
        lines.extend(f"- {table}: {count}" for table, count in sorted(removed_rows.items()))
    else:
        lines.append("- none")
    # 其余禁读类:即使已结构性隔离(run_load_dcf 只读沙箱 defaults / 只写 load forecast),
    # 也显式列入 AI 禁读清单,给模型理解阶段完整边界(SKILL §5 承诺 6 类)。
    agent = agent_dir(company_dir)
    root_core = sorted(company_dir.glob("*核心假设*.md"))
    lines.extend(["", "## Other forbidden materials (formal current-state artifacts)"])
    lines.append(f"- formal `{agent / 'defaults.yaml'}` (use sandbox defaults.yaml only)")
    lines.append(f"- formal `{agent / 'forecast'}` (load DCF writes only to Load forecast dir)")
    if root_core:
        lines.extend(f"- root core-assumption: {p}" for p in root_core)
    else:
        lines.append("- root core-assumption: none found")
    lines.append(f"- `{company_dir / 'WEBCLAUDE'}` (packaged copies, not source of truth)")
    return "\n".join(lines)


def _core_assumption_scaffold(company_dir: Path, boundary: ModelBoundary) -> str:
    name = company_dir.name.rsplit("_", 1)[0]
    horizon = ", ".join(str(year) for year in boundary.forecast_years)
    return f"""# {boundary.model_path.stem}_核心假设
> 模式: load | 模型源: {boundary.model_path.name} | 模型日期: {boundary.model_asof_date or "unknown"} | 历史止于 {boundary.history_end_year} | 显式预测 [{horizon}]
> 公司: {name}

## load 纪律

- 本文件用于保存原外部模型的 vintage 认知，不代表当前最新判断。
- 禁止用 {boundary.forecast_start_year} 年及之后的年报、季报、数据库实际值修正本模型。
- 后验已知错误也要保留为“原模型当时的预测”，只可在注释中标为 load-vintage。
- 本文件必须写成 `/comp` 可编译的核心假设源语言：业务线、上挂科目、compiler family、历史原子、预测旋钮、出处、收纳区和末尾 ```knobs 机器自报清单。

## 待模型装载器补全

按最新版 `skills/模型装载器_skill_v*.md` 读取 allowed_materials 中的模型公式层，补全业务线、历史原子、预测旋钮、公式关系和出处。
"""


def prepare_load(
    company: str | Path,
    *,
    model_path: str | Path | None = None,
    load_id: str | None = None,
    overwrite: bool = False,
    companies_dir: Path = COMPANIES_DIR,
) -> dict[str, Any]:
    company_dir = resolve_company(company, companies_dir=companies_dir)
    model = Path(model_path) if model_path else load_source_model(company_dir)
    if "WEBCLAUDE" in model.parts:
        raise ModelLoadError(f"/load refuses WEBCLAUDE packaged copies: {model}")

    boundary = inspect_workbook_boundary(model)
    if boundary.conflicts:
        raise ModelLoadError("; ".join(boundary.conflicts))

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", model.stem).strip("_") or "model"
    load_name = load_id or f"{safe_stem}_{boundary.history_end_year}_{stamp}"
    load_dir = agent_dir(company_dir) / "Load" / load_name
    if load_dir.exists():
        if not overwrite:
            raise ModelLoadError(f"load sandbox already exists: {load_dir}")
        shutil.rmtree(load_dir)
    load_dir.mkdir(parents=True)

    allowed_reports, forbidden_reports = _annual_materials(
        company_dir,
        history_end_year=boundary.history_end_year,
        forecast_start_year=boundary.forecast_start_year,
    )
    copied_materials = _copy_allowed_materials(load_dir, model, allowed_reports)

    removed_rows: dict[str, int] = {}
    source_db = company_db_path(company_dir)
    cutoff_db = load_dir / "data_cutoff.db"
    defaults = load_dir / "defaults.yaml"
    if source_db.exists():
        removed_rows = copy_cutoff_database(source_db, cutoff_db, boundary.history_end_year)
        defaults_data = build_defaults(cutoff_db)
        write_yaml2(defaults, defaults_data)

    boundary_json = load_dir / "model_boundary.json"
    boundary_json.write_text(
        json.dumps(boundary.as_dict(), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    _write_text(load_dir / "model_boundary.md", _boundary_markdown(boundary))
    _write_text(load_dir / "forbidden_materials.md", _forbidden_markdown(boundary, forbidden_reports, removed_rows, company_dir))
    # LOAD 参考稿统一落 KA 参考稿区（ka_reference_dir），命名 核心假设参考load_YYYYMMDD.md，
    # 与 brkd/alphapai 参考稿同处，/ka 到该目录找 核心假设参考*.md。沙箱副本同名保留在
    # load_dir，供 /load 编译 yaml1_load 与沙箱 DCF；两份内容必须一字不差。
    core_filename = f"核心假设参考load_{datetime.now().strftime('%Y%m%d')}.md"
    core_path = load_dir / core_filename
    root_core_path = ka_reference_dir(company_dir) / core_filename
    ka_reference_dir(company_dir).mkdir(parents=True, exist_ok=True)
    _write_text(core_path, _core_assumption_scaffold(company_dir, boundary))

    manifest = {
        "mode": "load_vintage",
        "company_dir": str(company_dir),
        "load_dir": str(load_dir),
        "model_path": str(model),
        "boundary": boundary.as_dict(),
        "allowed_materials": [str(path) for path in copied_materials],
        "forbidden_reports": [str(path) for path in forbidden_reports],
        "removed_rows": removed_rows,
        "data_cutoff_db": str(cutoff_db) if cutoff_db.exists() else None,
        "defaults_path": str(defaults) if defaults.exists() else None,
        "core_assumption_path": str(core_path),
        "core_assumption_scaffold_path": str(core_path),
        "root_core_assumption_path": str(root_core_path),
        "core_assumption_filename": core_filename,
    }
    (load_dir / "load_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return manifest


def _check_load_core_consistency(load_path: Path) -> None:
    """audit H5:/load 主产物(公司根目录)与沙箱副本是两份各自 LLM 手写的副本,无代码保同步。

    DCF 是 load 出结果的关口:此处比对两份内容,漂移即 raise——避免 `/ka` 读到的根目录主产物
    与本次 load DCF 实际依据的沙箱稿静默不一致(模型理解与估值脱节)。只有两份都存在且内容
    不同才拦;根稿尚未回填(LLM 还没写)则跳过,不阻塞沙箱内的独立试算。
    """
    manifest_path = load_path / "load_manifest.json"
    if not manifest_path.exists():
        return
    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return
    sandbox = manifest.get("core_assumption_path")
    root = manifest.get("root_core_assumption_path")
    if not sandbox or not root:
        return
    sp, rp = Path(sandbox), Path(root)
    if not (sp.exists() and rp.exists()):
        return

    def _norm(p: Path) -> str:
        return "\n".join(line.rstrip() for line in p.read_text(encoding="utf-8").splitlines()).strip()

    if _norm(sp) != _norm(rp):
        raise ModelLoadError(
            "load 主产物(公司根目录核心假设)与沙箱副本内容已漂移——/ka 读根稿、本次 load "
            "DCF 依据沙箱稿,两者不一致会让模型理解与估值脱节。请以一份为准同步另一份后重跑。\n"
            f"  根目录: {rp}\n  沙箱  : {sp}"
        )


def run_load_dcf(load_dir: str | Path, yaml1_path: str | Path) -> ForecastRun:
    load_path = Path(load_dir)
    defaults = load_path / "defaults.yaml"
    cutoff_db = load_path / "data_cutoff.db"
    if not defaults.exists():
        raise ModelLoadError(f"missing load defaults.yaml: {defaults}")
    if not cutoff_db.exists():
        raise ModelLoadError(f"missing load data_cutoff.db: {cutoff_db}")
    _check_load_core_consistency(load_path)
    return run_company_forecast(
        yaml1_path=Path(yaml1_path),
        defaults_path=defaults,
        clean_annual_path=cutoff_db,
        output_dir=load_path / "forecast",
        internal_dir=load_path / ".modelking",
        skip_staleness_gate=True,
        mode="load_vintage",
        write_derived_outputs=False,
    )


def _print_prepare_result(result: dict[str, Any]) -> None:
    boundary = result["boundary"]
    print(f"Load sandbox: {result['load_dir']}")
    print(f"Model: {result['model_path']}")
    print(f"History end: {boundary['history_end_year']}")
    print(f"Forecast start: {boundary['forecast_start_year']}")
    print(f"Core assumption scaffold: {result['core_assumption_path']}")
    print(f"Root core assumption output: {result['root_core_assumption_path']}")
    if result.get("defaults_path"):
        print(f"Defaults: {result['defaults_path']}")
    else:
        print("Defaults: skipped (missing Agent/data.db)")


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Prepare and run vintage /load model sandboxes.")
    sub = parser.add_subparsers(dest="command", required=True)

    prepare = sub.add_parser("prepare", help="Create a load sandbox from the latest active model")
    prepare.add_argument("company", help="Company name, code, ticker, or company directory")
    prepare.add_argument("--model", help="Explicit model workbook path")
    prepare.add_argument("--load-id", help="Deterministic load sandbox id")
    prepare.add_argument("--overwrite", action="store_true", help="Replace an existing sandbox with the same load id")
    prepare.add_argument("--json", action="store_true", help="Print the manifest JSON")

    dcf = sub.add_parser("dcf", help="Run sandbox DCF from a load yaml1")
    dcf.add_argument("--load-dir", required=True, help="Agent/Load/<load_id> directory")
    dcf.add_argument("--yaml1", required=True, help="yaml1_load*.yaml path")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    if args.command == "prepare":
        result = prepare_load(
            args.company,
            model_path=args.model,
            load_id=args.load_id,
            overwrite=args.overwrite,
        )
        if args.json:
            print(json.dumps(result, ensure_ascii=False, indent=2))
        else:
            _print_prepare_result(result)
        return 0
    if args.command == "dcf":
        run = run_load_dcf(args.load_dir, args.yaml1)
        print(f"Written load forecast: {run.output_dir}")
        print(f"Per-share value: {run.summary['per_share_value']}")
        if run.warnings_count:
            print(f"Warnings: {run.warnings_count} (details in load clean report)")
        return 0
    raise ModelLoadError(f"unknown command: {args.command}")


if __name__ == "__main__":
    raise SystemExit(main())
