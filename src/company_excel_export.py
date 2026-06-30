"""Export Boshi-style presentation workbooks from shared derived metrics."""

from __future__ import annotations

import csv
import json
import re
import shutil
import sqlite3
import tempfile
import zipfile
from copy import copy
from dataclasses import asdict
from datetime import date, datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import yaml
from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src import app_config
from src import field_registry as _registry
from src.company_paths import (
    forecast_dir as company_forecast_dir,
    db_path as company_db_path,
    latest_yaml1_path,
    official_breakdowns_dir,
    quarterly_reports_dir,
)
from src.annual_report_utils import parallel_map
from src.derived_metrics import DERIVED_METRICS_FILENAME


BASE_DIR = Path(__file__).resolve().parent.parent
DEFAULT_TEMPLATE_PATH = BASE_DIR / "templates" / "boshi_company_output.xlsm"
DEFAULT_OUTPUT_EXTENSION = ".xlsx"
DEFAULT_WORKBOOK_AUTHOR = "ModelKing"
COMMENT_SHEET_TITLES = {"\u70b9\u8bc4\u6a21\u677f", "\ub4d0\ud300\uce5c\uacbc"}
OOXML_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OOXML_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
OOXML_CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
MODEL_FONT = "Arial"
MODEL_NAVY = "1F1F1F"
MODEL_HEADER_BLUE = "3F4E5F"
MODEL_SECTION_BLUE = "E8EEF6"
MODEL_FORECAST_FILL = "EAF2F8"
MODEL_ASSUMPTION_FILL = "FFF2CC"
MODEL_ASSUMPTION_BORDER = "B7A25B"
MODEL_SUBTLE_FILL = "F7F9FB"
MODEL_GRID = "7F8FA6"
MODEL_LIGHT_GRID = "E3E7ED"
MODEL_GREY = "666666"
MODEL_BLUE_FONT = "0563C1"
MODEL_SECTION_FONT = "1F4E79"
FULL_STATEMENT_SHEETS = (
    ("full_is.csv", "完整利润表", "is", "利润表", "百万元"),
    ("full_bs.csv", "完整资产负债表", "bs", "资产负债表", "百万元"),
    ("full_cf.csv", "完整现金流量表", "cf", "现金流量表", "百万元"),
)
SEMIANNUAL_IS_SHEET = "半年度利润表"
SEMIANNUAL_REVENUE_SPLIT_SHEET = "半年度收入拆分"
QUARTERLY_IS_SHEET = "季度利润表"
QUARTERLY_STATE_BADGE: dict[str, tuple[str, str]] = {
    "actual": ("实", MODEL_GREY),
    "inherit": ("继", MODEL_GREY),
    "manual": ("人", "8B1E2D"),
    "q4": ("Q4", "D6A100"),
}
STATEMENT_KEY_ROWS: dict[str, set[str]] = {
    "is": {"revenue", "total_revenue", "operate_profit", "total_profit", "n_income_attr_p"},
    "bs": {"money_cap", "total_assets", "total_liab", "total_hldr_eqy_inc_min_int", "total_hldr_eqy_exc_min_int"},
    "cf": {"n_cashflow_act", "n_cashflow_inv_act", "n_cash_flows_fnc_act", "n_incr_cash_cash_equ", "c_cash_equ_end_period"},
}
STATEMENT_HIDDEN_ROWS: dict[str, set[str]] = {
    "is": {"total_opcost"},
}
INCOME_DERIVED_ROWS: dict[str, list[tuple[str, str, str]]] = {
    "revenue": [("revenue_yoy", "收入同比", "signed_percent")],
    "oper_cost": [("gross_margin", "毛利率", "percent")],
    "sell_exp": [("sell_exp_rate", "销售费用率", "percent")],
    "admin_exp": [("admin_exp_rate", "管理费用率", "percent")],
    "rd_exp": [("rd_exp_rate", "研发费用率", "percent")],
    "fin_exp": [("fin_exp_rate", "财务费用率", "percent")],
    "total_cogs": [("total_cogs_rate", "营业总成本率", "percent")],
    "operate_profit": [("operate_margin", "营业利润率", "percent")],
    "total_profit": [("total_profit_margin", "利润总额率", "percent")],
    "income_tax": [("effective_tax_rate", "所得税率", "percent")],
    "n_income_attr_p": [("n_income_attr_p_margin", "净利率", "percent"), ("n_income_attr_p_yoy", "净利润同比", "signed_percent")],
}
SEMIANNUAL_IS_SECTIONS: tuple[tuple[str, tuple[tuple[str, str, str, bool, bool], ...]], ...] = (
    (
        "收入与毛利",
        (
            ("revenue", "营业收入", "number", True, False),
            ("revenue_yoy", "同比增速", "signed_percent", False, True),
            ("oper_cost", "营业成本", "number", False, False),
            ("gross_profit", "毛利润", "number", True, False),
            ("gross_margin", "毛利率", "percent", False, True),
        ),
    ),
    (
        "期间费用",
        (
            ("biz_tax_surchg", "税金及附加", "number", False, False),
            ("sell_exp", "销售费用", "number", False, False),
            ("sell_exp_rate", "销售费用率", "percent", False, True),
            ("admin_exp", "管理费用", "number", False, False),
            ("admin_exp_rate", "管理费用率", "percent", False, True),
            ("rd_exp", "研发费用", "number", False, False),
            ("rd_exp_rate", "研发费用率", "percent", False, True),
            ("fin_exp", "财务费用", "number", False, False),
            ("fin_exp_rate", "财务费用率", "percent", False, True),
        ),
    ),
    (
        "利润",
        (
            ("operate_profit", "营业利润", "number", True, False),
            ("operate_margin", "营业利润率", "percent", False, True),
            ("total_profit", "利润总额", "number", True, False),
            ("income_tax", "所得税", "number", False, False),
            ("effective_tax_rate", "有效税率", "percent", False, True),
            ("n_income", "净利润", "number", False, False),
            ("n_income_margin", "净利率", "percent", False, True),
            ("n_income_attr_p", "归母净利润", "number", True, False),
            ("n_income_attr_p_yoy", "归母净利润同比", "signed_percent", False, True),
            ("n_income_attr_p_margin", "归母净利率", "percent", False, True),
        ),
    ),
)
SEMIANNUAL_AMOUNT_FIELDS = {
    "revenue",
    "oper_cost",
    "biz_tax_surchg",
    "sell_exp",
    "admin_exp",
    "rd_exp",
    "fin_exp",
    "operate_profit",
    "total_profit",
    "income_tax",
    "n_income",
    "n_income_attr_p",
}
REVENUE_SPLIT_SECTION_ORDER = ("整体收入", "产品", "地区", "行业", "销售模式")
ASSUMPTION_SECTION_DEFS = (
    ("毛利率", ("income.gpm",)),
    ("费用率", ("income.cost_rates.", "income.financial_expense.other_fin_exp_abs")),
    ("营业利润调节 / 营业外收支（绝对值）", ("income.cost_abs.", "income.operating_adjustments_abs.", "income.below_line_abs.")),
    ("税率 / 少数股东", ("income.effective_tax_rate", "income.minority_ratio")),
)
EXPENSE_SECTION_ORDER = {
    "income.cost_rates.sell_exp": 0,
    "income.cost_rates.admin_exp": 1,
    "income.cost_rates.rd_exp": 2,
    "income.cost_rates.biz_tax_surchg": 3,
    "income.financial_expense.other_fin_exp_abs": 999,
}
ASSUMPTION_LABELS = {
    "gpm": "整体毛利率",
    "sell_exp": "销售费用率",
    "admin_exp": "管理费用率",
    "rd_exp": "研发费用率",
    "biz_tax_surchg": "税金及附加率",
    "other_fin_exp_abs": "非息财务费用",
    "effective_tax_rate": "有效税率",
    "minority_ratio": "少数股东损益占比",
    "assets_impair_loss": "资产减值损失",
    "credit_impa_loss": "信用减值损失",
    "oth_income": "其他收益",
    "invest_income": "投资收益",
    "fv_value_chg_gain": "公允价值变动收益",
    "asset_disp_income": "资产处置收益",
    "non_oper_income": "营业外收入",
    "non_oper_exp": "营业外支出",
    "revenue_yoy": "收入增长",
    "volume": "销量增长",
    "price": "单价增长",
}

SUMMARY_COLUMNS = tuple(get_column_letter(col) for col in range(65, 79))  # BM:BZ
SUMMARY_YEARS = tuple(range(2015, 2029))
RATING_COLUMNS = tuple(get_column_letter(col) for col in range(2, 10))  # B:I
COMMENT_RATING_COLUMNS = tuple(get_column_letter(col) for col in range(3, 11))  # C:J
COMMENT_ANNUAL_COLUMNS = tuple(get_column_letter(col) for col in range(3, 11))  # C:J
COMMENT_QUARTERLY_COLUMNS = tuple(get_column_letter(col) for col in range(3, 19))  # C:R

SUMMARY_ROW_MAP: dict[int, str] = {
    10: "revenue",
    11: "revenue_yoy",
    12: "oper_cost",
    13: "gross_profit",
    14: "biz_tax_surchg",
    15: "biz_tax_surchg_rate",
    16: "sales_profit",
    17: "gross_margin",
    18: "sell_exp",
    19: "sell_exp_rate",
    20: "admin_exp",
    21: "admin_exp_rate",
    22: "impairment",
    23: "impairment_rate",
    24: "sgna",
    25: "sgna_rate",
    26: "ebitda",
    27: "ebitda_margin",
    28: "da",
    29: "ebit",
    30: "ebit_margin",
    31: "fin_exp",
    32: "invest_income_fv",
    33: "non_operating_net",
    34: "total_profit",
    35: "income_tax",
    36: "effective_tax_rate",
    37: "minority_gain",
    38: "minority_gain_rate",
    39: "n_income_attr_p",
    40: "n_income_attr_p_margin",
    41: "n_income_attr_p_yoy",
    44: "fixed_intangible_longterm_assets",
    45: "operating_wc_assets",
    46: "operating_wc_liabilities",
    47: "operating_nwc",
    48: "invested_capital",
    49: "invested_capital_turnover",
    50: "cash",
    51: "interest_bearing_debt",
    52: "net_cash",
    53: "minority_int",
    54: "parent_equity",
    55: "total_assets",
    59: "cfo",
    60: "capex",
    61: "investment_acquisition",
    62: "cfi",
    63: "equity_financing",
    64: "cff",
    65: "debt_financing",
    66: "cash_net_change",
    67: "net_cash_change",
    69: "fcf",
    72: "total_shares",
    73: "eps",
    76: "bvps",
    78: "pe",
    79: "pb",
    81: "market_cap",
    82: "avg_minority_int",
    83: "avg_net_debt",
    84: "enterprise_value",
    85: "ev_ebitda",
    86: "ev_sales",
    87: "tax_burden",
    88: "interest_burden",
    89: "sales_profit_margin",
    90: "asset_turnover",
    91: "roa",
    92: "leverage",
    93: "roe",
    94: "roic",
    95: "capex_to_revenue",
    96: "capex_to_ebitda",
    97: "capex_to_da",
    98: "asset_liability_ratio",
    99: "net_debt_ratio",
    100: "ebitda_interest_coverage",
}

RATING_ROW_MAP: dict[int, str] = {
    2: "revenue",
    3: "revenue_yoy",
    4: "gross_margin",
    5: "n_income_attr_p",
    6: "n_income_attr_p_yoy",
    7: "roe",
    8: "eps",
    9: "pe",
    10: "pb",
    11: "ev_ebitda",
}

COMMENT_ANNUAL_ROW_MAP: dict[int, str] = {
    4: "revenue",
    5: "revenue_yoy",
    6: "oper_cost",
    7: "oper_cost_yoy",
    8: "n_income_attr_p",
    9: "n_income_attr_p_yoy",
    10: "n_income_attr_p",
    11: "n_income_attr_p_yoy",
    12: "gross_margin",
    13: "sell_exp_rate",
    14: "gross_sell_spread",
    15: "admin_exp_rate",
    16: "rd_exp_rate",
    17: "fin_exp_rate",
    18: "effective_tax_rate",
    20: "n_income_attr_p_margin",
    21: "n_income_attr_p_margin",
}

COMMENT_QUARTERLY_ROW_MAP: dict[int, str] = {
    31: "revenue",
    32: "revenue_yoy",
    33: "oper_cost",
    34: "oper_cost_yoy",
    35: "n_income_attr_p",
    36: "n_income_attr_p_yoy",
    37: "n_income_attr_p",
    38: "n_income_attr_p_yoy",
    40: "gross_margin",
    41: "sell_exp_rate",
    42: "gross_sell_spread",
    43: "admin_exp_rate",
    44: "rd_exp_rate",
    45: "fin_exp_rate",
    46: "effective_tax_rate",
    48: "n_income_attr_p_margin",
    49: "n_income_attr_p_margin",
}

MODEL_BS_ROW_MAP: dict[int, tuple[str, str]] = {
    5: ("metric", "revenue"),
    6: ("metric", "revenue_yoy"),
    7: ("metric", "oper_cost"),
    8: ("calc", "oper_cost_rate"),
    9: ("metric", "biz_tax_surchg"),
    10: ("metric", "biz_tax_surchg_rate"),
    11: ("metric", "sales_profit"),
    12: ("metric", "sales_profit_margin"),
    14: ("metric", "sell_exp"),
    15: ("metric", "sell_exp_rate"),
    16: ("metric", "admin_exp"),
    17: ("metric", "admin_exp_rate"),
    18: ("metric", "rd_exp"),
    19: ("metric", "rd_exp_rate"),
    20: ("metric", "impairment"),
    21: ("metric", "impairment_rate"),
    22: ("is", "assets_impair_loss"),
    23: ("is", "credit_impa_loss"),
    24: ("metric", "sgna"),
    25: ("metric", "sgna_rate"),
    27: ("metric", "ebitda"),
    28: ("metric", "ebitda_margin"),
    29: ("metric", "da"),
    31: ("metric", "ebit"),
    32: ("metric", "ebit_margin"),
    34: ("metric", "fin_exp"),
    35: ("is", "fin_exp_int_inc"),
    36: ("is", "fin_exp_int_exp"),
    37: ("calc", "other_fin_exp"),
    41: ("is", "oth_income"),
    42: ("is", "invest_income"),
    45: ("is", "fv_value_chg_gain"),
    46: ("is", "asset_disp_income"),
    47: ("is", "forex_gain"),
    49: ("is", "non_oper_income"),
    50: ("is", "non_oper_exp"),
    51: ("metric", "non_operating_net"),
    52: ("calc", "income_adjustments_total"),
    54: ("metric", "total_profit"),
    55: ("metric", "income_tax"),
    56: ("metric", "effective_tax_rate"),
    57: ("metric", "minority_gain"),
    58: ("metric", "minority_gain_rate"),
    59: ("metric", "n_income_attr_p"),
    60: ("metric", "n_income_attr_p_margin"),
    61: ("metric", "n_income_attr_p_yoy"),
    63: ("calc", "zero"),
    64: ("metric", "total_shares"),
    120: ("metric", "eps"),
    121: ("metric", "eps"),
    122: ("metric", "dps"),
    124: ("calc", "dividend_total"),
    130: ("metric", "total_assets"),
    131: ("metric", "total_liab"),
    132: ("metric", "total_equity"),
    134: ("calc", "balance_check"),
    138: ("bs", "total_cur_assets"),
    140: ("metric", "cash"),
    141: ("bs", "oth_cur_assets"),
    143: ("metric", "operating_wc_assets"),
    144: ("bs", "inventories"),
    145: ("calc", "inventories_rate"),
    146: ("calc", "receivables_total"),
    147: ("calc", "receivables_rate"),
    148: ("bs", "notes_receiv"),
    149: ("bs", "accounts_receiv"),
    150: ("bs", "prepayment"),
    155: ("bs", "oth_receiv"),
    157: ("bs", "contract_assets"),
    159: ("bs", "total_nca"),
    161: ("bs", "goodwill"),
    162: ("bs", "oth_nca"),
    166: ("bs", "fix_assets"),
    177: ("bs", "cip"),
    179: ("bs", "intan_assets"),
    190: ("metric", "da"),
    191: ("metric", "capex"),
    193: ("bs", "total_cur_liab"),
    195: ("bs", "st_borr"),
    197: ("calc", "non_operating_current_liab"),
    199: ("metric", "operating_wc_liabilities"),
    200: ("calc", "operating_wc_liabilities_rate"),
    201: ("bs", "notes_payable"),
    202: ("bs", "acct_payable"),
    203: ("bs", "adv_receipts"),
    204: ("bs", "contract_liab"),
    206: ("bs", "payroll_payable"),
    207: ("bs", "taxes_payable"),
    208: ("bs", "int_payable"),
    209: ("bs", "div_payable"),
    210: ("bs", "oth_payable"),
    213: ("bs", "total_ncl"),
    215: ("calc", "interest_bearing_ncl"),
    216: ("bs", "lt_borr"),
    217: ("bs", "bond_payable"),
    219: ("calc", "non_interest_ncl"),
    221: ("metric", "interest_bearing_debt"),
    222: ("is", "fin_exp_int_exp"),
    223: ("calc", "interest_expense_rate"),
    225: ("metric", "cash"),
    226: ("is", "fin_exp_int_inc"),
    227: ("calc", "interest_income_rate"),
    229: ("metric", "total_equity"),
    230: ("bs", "total_share"),
    232: ("bs", "cap_rese"),
    233: ("bs", "treasury_share"),
    234: ("bs", "oth_comp_income"),
    235: ("bs", "surplus_rese"),
    236: ("bs", "ordin_risk_reser"),
    237: ("bs", "undistr_porfit"),
    239: ("bs", "invest_loss_unconf"),
    240: ("bs", "minority_int"),
    243: ("calc", "zero"),
    246: ("metric", "operating_nwc"),
    247: ("metric", "operating_wc_assets"),
    248: ("metric", "operating_wc_liabilities"),
    250: ("calc", "operating_assets_decrease"),
    251: ("calc", "operating_liabilities_increase"),
    253: ("bs", "defer_tax_assets"),
    254: ("calc", "defer_tax_assets_decrease"),
    255: ("bs", "defer_tax_liab"),
    256: ("calc", "defer_tax_liab_increase"),
    259: ("metric", "n_income_attr_p"),
    260: ("metric", "minority_gain"),
    261: ("is", "assets_impair_loss"),
    262: ("is", "credit_impa_loss"),
    263: ("metric", "da"),
    264: ("metric", "fin_exp"),
    265: ("is", "invest_income"),
    266: ("is", "fv_value_chg_gain"),
    267: ("calc", "operating_assets_decrease"),
    268: ("calc", "operating_liabilities_increase"),
    269: ("calc", "defer_tax_assets_decrease"),
    270: ("calc", "defer_tax_liab_increase"),
    273: ("metric", "cfo"),
    274: ("metric", "cfo"),
    275: ("metric", "cfo"),
    277: ("metric", "capex"),
    278: ("metric", "investment_acquisition"),
    284: ("metric", "cfi"),
    287: ("metric", "equity_financing"),
    292: ("cf", "c_recp_borrow"),
    294: ("cf", "c_prepay_amt_borr"),
    295: ("cf", "c_pay_dist_dpcp_int_exp"),
    296: ("is", "fin_exp_int_exp"),
    297: ("calc", "dividend_total"),
    299: ("metric", "cff"),
    302: ("cf", "eff_fx_flu_cash"),
    303: ("metric", "cash_net_change"),
    306: ("cf", "c_cash_equ_beg_period"),
    307: ("cf", "c_cash_equ_end_period"),
    311: ("cf", "c_cash_equ_beg_period"),
    312: ("cf", "c_cash_equ_end_period"),
    313: ("metric", "cash_net_change"),
    315: ("calc", "balance_check"),
    321: ("metric", "ebit"),
    322: ("metric", "ebitda"),
    323: ("calc", "nopat"),
    324: ("metric", "n_income_attr_p"),
    326: ("metric", "gross_margin"),
    327: ("metric", "ebit_margin"),
    328: ("metric", "ebitda_margin"),
    329: ("metric", "n_income_attr_p_margin"),
    332: ("metric", "market_cap"),
    333: ("metric", "enterprise_value"),
    334: ("metric", "invested_capital"),
    335: ("metric", "operating_nwc"),
    338: ("metric", "roa"),
    339: ("metric", "roe"),
    340: ("metric", "roic"),
    343: ("metric", "pb"),
    344: ("metric", "dividend_yield"),
    345: ("metric", "pe"),
    346: ("metric", "ev_ebitda"),
    347: ("metric", "eps"),
    351: ("calc", "nopat"),
    352: ("metric", "da"),
    353: ("metric", "capex"),
    354: ("calc", "delta_nwc"),
    355: ("calc", "fcff"),
    358: ("calc", "cash_collection_quality"),
    359: ("calc", "cash_profit_quality"),
    360: ("calc", "contract_liab_change_to_revenue"),
    361: ("calc", "inventory_change_to_cogs"),
}


def _load_metrics(company_dir: Path) -> dict[str, Any]:
    path = company_forecast_dir(company_dir) / DERIVED_METRICS_FILENAME
    if not path.exists():
        raise FileNotFoundError(f"derived metrics not found: {path}")
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    if not isinstance(data, dict):
        raise ValueError(f"derived metrics root must be an object: {path}")
    return data


def _safe_filename_part(value: Any, fallback: str = "Company") -> str:
    text = str(value or fallback).strip()
    text = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "", text)
    text = text.strip(" .")
    return text or fallback


def _workbook_author(researcher_name: str) -> str:
    return researcher_name.strip() or DEFAULT_WORKBOOK_AUTHOR


def _apply_output_identity(workbook: Any, researcher_name: str) -> None:
    author = _workbook_author(researcher_name)
    workbook.properties.creator = author
    workbook.properties.lastModifiedBy = author


def _default_output_path(company_dir: Path, metrics: dict[str, Any]) -> Path:
    name = metrics.get("name") or metrics.get("ticker") or company_dir.name.split("_", 1)[0]
    date_code = date.today().strftime("%y%m%d")
    return company_dir / f"{_safe_filename_part(name)}Model-{date_code}{DEFAULT_OUTPUT_EXTENSION}"


def _numbered_output_path(path: Path) -> Path:
    for index in range(1, 100):
        candidate = path.with_name(f"{path.stem}-{index}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise FileExistsError(f"No available numbered output path for {path}")


def _sheet(workbook: Any, title: str, index: int) -> Any:
    if title in workbook.sheetnames:
        return workbook[title]
    return workbook.worksheets[index]


def _remove_comment_sheets(workbook: Any) -> None:
    targets = [ws for ws in workbook.worksheets if ws.title in COMMENT_SHEET_TITLES]
    if not targets and len(workbook.worksheets) >= 4:
        targets = [workbook.worksheets[2]]
    for ws in targets:
        workbook.remove(ws)


def _xml_bytes(root: ET.Element) -> bytes:
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _drop_external_relationships(data: bytes) -> bytes:
    root = ET.fromstring(data)
    for rel in list(root):
        rel_type = rel.attrib.get("Type", "")
        target = rel.attrib.get("Target", "")
        if rel_type.endswith("/externalLink") or "externalLinks/" in target:
            root.remove(rel)
    return _xml_bytes(root)


def _drop_workbook_problem_nodes(data: bytes) -> bytes:
    root = ET.fromstring(data)
    for child in list(root):
        if child.tag in {f"{{{OOXML_MAIN_NS}}}externalReferences", f"{{{OOXML_MAIN_NS}}}definedNames"}:
            root.remove(child)
        elif child.tag.endswith("externalReferences") or child.tag.endswith("definedNames"):
            root.remove(child)
    return _xml_bytes(root)


def _drop_external_content_types(data: bytes) -> bytes:
    root = ET.fromstring(data)
    for item in list(root):
        part_name = item.attrib.get("PartName", "")
        content_type = item.attrib.get("ContentType", "")
        if part_name.startswith("/xl/externalLinks/") or "externalLink" in content_type:
            root.remove(item)
    return _xml_bytes(root)


def _sanitize_workbook_package(package_path: Path) -> None:
    """Remove stale workbook parts that Excel may try to repair."""
    with tempfile.TemporaryDirectory(prefix="boshi_excel_links_") as tmp_dir:
        cleaned_path = Path(tmp_dir) / package_path.name
        with zipfile.ZipFile(package_path, "r") as source, zipfile.ZipFile(cleaned_path, "w", zipfile.ZIP_DEFLATED) as target:
            for info in source.infolist():
                name = info.filename.replace("\\", "/")
                if name.startswith("xl/externalLinks/"):
                    continue
                data = source.read(info.filename)
                if name == "xl/_rels/workbook.xml.rels":
                    data = _drop_external_relationships(data)
                elif name == "xl/workbook.xml":
                    data = _drop_workbook_problem_nodes(data)
                elif name == "[Content_Types].xml":
                    data = _drop_external_content_types(data)
                target.writestr(info, data)
        package_path.unlink()
        shutil.move(cleaned_path, package_path)


def _fresh_sheet(workbook: Any, title: str) -> Any:
    if title in workbook.sheetnames:
        workbook.remove(workbook[title])
    return workbook.create_sheet(title=title)


def _style_report_sheet(ws: Any, *, title: str, subtitle: str, max_col: int, frozen_col: int = 2) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"{get_column_letter(frozen_col)}4"
    ws.sheet_properties.tabColor = MODEL_NAVY
    ws.sheet_view.zoomScale = 90
    ws.sheet_format.defaultRowHeight = 16
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"] = title
    ws["A1"].font = Font(name=MODEL_FONT, bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = PatternFill("solid", fgColor=MODEL_NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name=MODEL_FONT, color=MODEL_GREY, italic=True, size=9)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 17
    ws.row_dimensions[3].height = 18
    ws.column_dimensions["A"].width = 32
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    thin = Side(style="thin", color=MODEL_LIGHT_GRID)
    ws._report_border = Border(bottom=thin)  # type: ignore[attr-defined]


def _header_style(cell: Any) -> None:
    cell.font = Font(name=MODEL_FONT, bold=True, color="FFFFFF", size=9)
    cell.fill = PatternFill("solid", fgColor=MODEL_HEADER_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(top=Side(style="thin", color=MODEL_NAVY), bottom=Side(style="medium", color=MODEL_NAVY))


def _group_style(cell: Any) -> None:
    cell.font = Font(name=MODEL_FONT, bold=True, color=MODEL_SECTION_FONT, size=9)
    cell.fill = PatternFill("solid", fgColor=MODEL_SECTION_BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = Border(top=Side(style="medium", color=MODEL_GRID), bottom=Side(style="thin", color=MODEL_GRID))


def _subsection_style(cell: Any) -> None:
    cell.font = Font(name=MODEL_FONT, bold=True, color=MODEL_GREY, size=9)
    cell.fill = PatternFill("solid", fgColor="FFFFFF")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = Border(top=Side(style="thin", color=MODEL_LIGHT_GRID), bottom=Side(style="thin", color=MODEL_LIGHT_GRID))


def _data_style(cell: Any, *, forecast: bool, number_format: str = '#,##0.0;[Red](#,##0.0);"-"') -> None:
    cell.number_format = number_format
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.font = Font(name=MODEL_FONT, color=MODEL_BLUE_FONT if forecast else "000000", size=9)
    cell.fill = PatternFill("solid", fgColor=MODEL_FORECAST_FILL if forecast else "FFFFFF")
    cell.border = Border(bottom=Side(style="thin", color=MODEL_LIGHT_GRID))


def _assumption_data_style(cell: Any, *, number_format: str = '#,##0.0;[Red](#,##0.0);"-"') -> None:
    cell.number_format = number_format
    cell.font = Font(name=MODEL_FONT, color=MODEL_BLUE_FONT, bold=True, size=9)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.fill = PatternFill("solid", fgColor=MODEL_ASSUMPTION_FILL)
    cell.border = Border(
        top=Side(style="thin", color=MODEL_ASSUMPTION_BORDER),
        bottom=Side(style="thin", color=MODEL_ASSUMPTION_BORDER),
    )


def _label_style(cell: Any, *, bold: bool = False, muted: bool = False) -> None:
    cell.font = Font(name=MODEL_FONT, bold=bold, italic=muted, color=MODEL_GREY if muted else "000000", size=9)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1 if muted else 0)
    cell.fill = PatternFill("solid", fgColor=MODEL_SUBTLE_FILL if bold else "FFFFFF")
    cell.border = Border(bottom=Side(style="thin", color=MODEL_LIGHT_GRID))


def _assumption_label_style(cell: Any, *, muted: bool = False) -> None:
    cell.font = Font(name=MODEL_FONT, bold=True, italic=muted, color=MODEL_BLUE_FONT, size=9)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1 if muted else 0)
    cell.fill = PatternFill("solid", fgColor=MODEL_ASSUMPTION_FILL)
    cell.border = Border(
        top=Side(style="thin", color=MODEL_ASSUMPTION_BORDER),
        bottom=Side(style="thin", color=MODEL_ASSUMPTION_BORDER),
    )


def _number_format(format_key: str) -> str:
    if format_key in {"percent", "signed_percent", "decimal", "decimal1"}:
        return '0.0%;[Red](0.0%);"-"'
    if format_key == "integer":
        return '#,##0;[Red](#,##0);"-"'
    if format_key == "multiple":
        return '0.0x;[Red](0.0x);"-"'
    return '#,##0.0;[Red](#,##0.0);"-"'


def _write_header(ws: Any, row: int, labels: list[str]) -> None:
    for col, label in enumerate(labels, 1):
        cell = ws.cell(row, col, label)
        _header_style(cell)
        if col == 1:
            cell.alignment = Alignment(horizontal="left", vertical="center")


def _auto_widths(ws: Any, max_col: int, *, start_col: int = 1) -> None:
    for col in range(start_col, max_col + 1):
        letter = get_column_letter(col)
        if col >= 2:
            ws.column_dimensions[letter].width = 12
        else:
            current = ws.column_dimensions[letter].width or 12
            ws.column_dimensions[letter].width = max(current, 34)


def _emphasize_row(ws: Any, row: int, max_col: int, *, forecast_start_col: int = 2) -> None:
    border = Border(top=Side(style="thin", color=MODEL_GRID), bottom=Side(style="medium", color=MODEL_GRID))
    for col in range(1, max_col + 1):
        cell = ws.cell(row, col)
        existing_color = cell.font.color.rgb if getattr(cell.font.color, "type", None) == "rgb" else None
        cell.font = Font(name=MODEL_FONT, bold=True, color=existing_color or ("000000" if col < forecast_start_col else MODEL_BLUE_FONT), size=9)
        if col < forecast_start_col:
            cell.fill = PatternFill("solid", fgColor=MODEL_SUBTLE_FILL)
        cell.border = border


def _emphasize_quarterly_row(
    ws: Any, row: int, periods: list[str], annual_col: int, selected_year: int
) -> None:
    """Bold a key row on the quarterly sheet (3-segment coloring)."""
    border = Border(top=Side(style="thin", color=MODEL_GRID), bottom=Side(style="medium", color=MODEL_GRID))
    ws.cell(row, 1).border = border
    for idx, period in enumerate(periods):
        col = 2 + idx
        forecast = int(period[:4]) == selected_year
        cell = ws.cell(row, col)
        cell.font = Font(name=MODEL_FONT, bold=True, color=MODEL_BLUE_FONT if forecast else "000000", size=9)
        if not forecast:
            cell.fill = PatternFill("solid", fgColor=MODEL_SUBTLE_FILL)
        cell.border = border
    annual_cell = ws.cell(row, annual_col)
    annual_cell.font = Font(name=MODEL_FONT, bold=True, color="000000", size=9)
    annual_cell.fill = PatternFill("solid", fgColor=MODEL_SUBTLE_FILL)
    annual_cell.border = border


def _load_yaml1(company_dir: Path) -> dict[str, Any]:
    try:
        path = latest_yaml1_path(company_dir)
    except FileNotFoundError:
        return {}
    try:
        data = yaml.safe_load(path.read_text(encoding="utf-8-sig"))
    except (OSError, yaml.YAMLError):
        return {}
    return data if isinstance(data, dict) else {}


def _yaml1_years(data: dict[str, Any]) -> list[str]:
    meta = data.get("meta")
    horizon = meta.get("horizon") if isinstance(meta, dict) else None
    if isinstance(horizon, list):
        years = []
        for item in horizon:
            try:
                years.append(str(int(item)))
            except (TypeError, ValueError):
                continue
        if years:
            return years
    max_len = 0
    for payload in data.values():
        if isinstance(payload, dict) and isinstance(payload.get("values"), list):
            max_len = max(max_len, len(payload["values"]))
    base = _yaml1_base_period(data)
    return [str(base + i) for i in range(1, max_len + 1)] if base and max_len else []


def _yaml1_base_period(data: dict[str, Any]) -> int | None:
    years = _yaml1_years_from_meta(data)
    return int(years[0]) - 1 if years else None


def _yaml1_years_from_meta(data: dict[str, Any]) -> list[str]:
    meta = data.get("meta")
    horizon = meta.get("horizon") if isinstance(meta, dict) else None
    if not isinstance(horizon, list):
        return []
    years = []
    for item in horizon:
        try:
            years.append(str(int(item)))
        except (TypeError, ValueError):
            continue
    return years


def _assumption_label(path: str) -> str:
    leaf = path.rsplit(".", 1)[-1]
    return ASSUMPTION_LABELS.get(leaf, _registry.field_label(leaf, leaf))


def _historical_assumption_value(
    metrics: dict[str, Any],
    statements: dict[str, dict[str, dict[str, Any]]],
    path: str,
    year: str,
) -> Any:
    revenue = _metric_value(metrics, year, "revenue")
    if path == "income.gpm":
        return _metric_value(metrics, year, "gross_margin")
    if path.startswith("income.cost_rates."):
        field = path.rsplit(".", 1)[-1]
        return _safe_div(_statement_value(statements, "is", year, field), revenue)
    if path == "income.effective_tax_rate":
        return _metric_value(metrics, year, "effective_tax_rate")
    if path == "income.minority_ratio":
        return _metric_value(metrics, year, "minority_gain_rate")
    if path.startswith("income."):
        field = path.rsplit(".", 1)[-1]
        return _statement_value(statements, "is", year, field)
    return None


def _row_has_value(values: dict[str, Any]) -> bool:
    return any(_num(value) is not None and abs(float(value)) > 1e-9 for value in values.values())


def _series_values_from_history(node: dict[str, Any], key: str) -> dict[str, float | None]:
    history = node.get("history")
    series = history.get("series") if isinstance(history, dict) else None
    values = series.get(key) if isinstance(series, dict) else None
    if not isinstance(values, dict):
        return {}
    return {str(year): _num(value) for year, value in values.items()}


def _growth_series(base: Any, yoy_values: list[Any], years: list[str]) -> dict[str, float | None]:
    current = _num(base)
    out: dict[str, float | None] = {}
    for year, yoy in zip(years, yoy_values):
        growth = _num(yoy)
        if current is None or growth is None:
            out[year] = None
            current = None
        else:
            current *= 1 + growth
            out[year] = current
    return out


def _growth_rate(current: Any, previous: Any) -> float | None:
    cur = _num(current)
    prev = _num(previous)
    if cur is None or prev is None:
        return None
    denominator = abs(prev) if prev < 0 else prev
    if abs(denominator) < 1e-9:
        return None
    return (cur - prev) / denominator


def _yoy_for_years(values: dict[str, Any], years: list[str]) -> dict[str, float | None]:
    out: dict[str, float | None] = {}
    previous: float | None = None
    for year in years:
        current = _num(values.get(year))
        out[year] = _growth_rate(current, previous)
        previous = current if current is not None else previous
    return out


def _segment_projection(node: dict[str, Any], years: list[str]) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any], dict[str, Any]]:
    base = node.get("base") if isinstance(node.get("base"), dict) else {}
    revenue_history = _series_values_from_history(node, "revenue")
    volume_history = _series_values_from_history(node, "volume")
    revenue_forecast: dict[str, Any] = {}
    yoy_forecast: dict[str, Any] = {}
    volume_forecast: dict[str, Any] = {}
    driver_rows: dict[str, dict[str, Any]] = {}
    family = str(node.get("revenue_family") or "")
    if family == "factor_product":
        factors = {str(item.get("key")): item for item in node.get("factors", []) if isinstance(item, dict)}
        volume_factor = factors.get("volume")
        price_factor = factors.get("price")
        volume_series = _growth_series(
            volume_factor.get("base") if isinstance(volume_factor, dict) else None,
            (volume_factor.get("projection") or {}).get("values", []) if isinstance(volume_factor, dict) else [],
            years,
        )
        price_series = _growth_series(
            price_factor.get("base") if isinstance(price_factor, dict) else None,
            (price_factor.get("projection") or {}).get("values", []) if isinstance(price_factor, dict) else [],
            years,
        )
        factor = _num(base.get("unit_factor_to_million_cny")) or 1
        for year in years:
            volume = volume_series.get(year)
            price = price_series.get(year)
            revenue_forecast[year] = None if volume is None or price is None else volume * price / factor
            volume_forecast[year] = volume
        for key, factor_node in factors.items():
            projection = factor_node.get("projection") if isinstance(factor_node.get("projection"), dict) else {}
            values = projection.get("values") if isinstance(projection.get("values"), list) else []
            driver_rows[key] = {year: _num(value) for year, value in zip(years, values)}
    else:
        yoy_values = (node.get("knobs") or {}).get("revenue_yoy") if isinstance(node.get("knobs"), dict) else None
        yoy_values = yoy_values if isinstance(yoy_values, list) else []
        current = _num(base.get("revenue"))
        for year, yoy in zip(years, yoy_values):
            growth = _num(yoy)
            if current is None or growth is None:
                revenue_forecast[year] = None
            else:
                current *= 1 + growth
                revenue_forecast[year] = current
            driver_rows["revenue_yoy"] = {year_item: _num(value) for year_item, value in zip(years, yoy_values)}
    previous = _num(base.get("revenue")) or _num(revenue_history.get(str(base.get("base_year"))))
    for year in years:
        current = _num(revenue_forecast.get(year))
        yoy_forecast[year] = _growth_rate(current, previous)
        previous = current if current is not None else previous
    return revenue_forecast, yoy_forecast, volume_forecast, driver_rows


def _core_assumption_years(metrics: dict[str, Any], yaml_years: list[str], base_period: int | None) -> list[str]:
    annual = metrics.get("annual") if isinstance(metrics.get("annual"), dict) else {}
    history = []
    if base_period:
        history = [year for year in annual if year.isdigit() and int(year) <= base_period]
    return sorted(set(history[-5:] + yaml_years), key=lambda year: int(year))


def _write_assumption_group(
    ws: Any,
    row: int,
    *,
    title: str,
    years: list[str],
    rows: list[dict[str, Any]],
    base_period: int | None,
) -> int:
    max_col = len(years) + 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row, 1, title)
    _group_style(cell)
    row += 1
    for item in rows:
        is_assumption = bool(item.get("assumption"))
        ws.cell(row, 1, item.get("label"))
        for col_offset, year in enumerate(years, 2):
            value = item.get("values", {}).get(year)
            cell = ws.cell(row, col_offset, value)
            forecast = bool(base_period and int(year) > base_period)
            if is_assumption and forecast:
                _assumption_data_style(cell, number_format=_number_format(item.get("format", "number")))
            else:
                _data_style(cell, forecast=forecast, number_format=_number_format(item.get("format", "number")))
        if is_assumption:
            _assumption_label_style(ws.cell(row, 1), muted=bool(item.get("muted")))
        else:
            _label_style(ws.cell(row, 1), bold=bool(item.get("bold")), muted=bool(item.get("muted")))
        if item.get("bold") and not is_assumption:
            _emphasize_row(ws, row, max_col)
        row += 1
    return row + 1


def _fill_core_assumptions_sheet(
    workbook: Any,
    company_dir: Path,
    metrics: dict[str, Any],
    statements: dict[str, dict[str, dict[str, Any]]],
) -> None:
    yaml1 = _load_yaml1(company_dir)
    if not yaml1:
        return
    yaml_years = _yaml1_years(yaml1)
    base_period = _yaml1_base_period(yaml1) or int(float(metrics.get("base_period") or 0)) or None
    years = _core_assumption_years(metrics, yaml_years, base_period)
    if not years:
        return
    ws = _fresh_sheet(workbook, "核心假设")
    _style_report_sheet(ws, title="核心假设", subtitle="单位：百万元；比率为百分比", max_col=len(years) + 1)
    _write_header(ws, 3, ["项目", *[f"{year}E" if base_period and int(year) > base_period else year for year in years]])
    row = 4
    total_rows = [
        {"label": "营业收入", "values": {year: _metric_value(metrics, year, "revenue") for year in years}, "format": "number", "bold": True},
        {"label": "同比增长", "values": {year: _metric_value(metrics, year, "revenue_yoy") for year in years}, "format": "signed_percent", "muted": True},
        {"label": "归母净利润", "values": {year: _metric_value(metrics, year, "n_income_attr_p") for year in years}, "format": "number", "bold": True},
        {"label": "净利润", "values": {year: _metric_value(metrics, year, "n_income") for year in years}, "format": "number"},
        {"label": "净利率", "values": {year: _metric_value(metrics, year, "n_income_attr_p_margin") for year in years}, "format": "percent", "muted": True},
        {"label": "净利润同比", "values": {year: _metric_value(metrics, year, "n_income_attr_p_yoy") for year in years}, "format": "signed_percent", "muted": True},
    ]
    row = _write_assumption_group(ws, row, title="总收入与利润路径", years=years, rows=total_rows, base_period=base_period)

    revenue_node = yaml1.get("income.revenue") if isinstance(yaml1.get("income.revenue"), dict) else {}
    segment_rows: list[dict[str, Any]] = []
    segments = revenue_node.get("segments") if isinstance(revenue_node, dict) else None
    if isinstance(segments, dict):
        for segment_name, node in segments.items():
            if not isinstance(node, dict):
                continue
            revenue_history = _series_values_from_history(node, "revenue")
            volume_history = _series_values_from_history(node, "volume")
            revenue_forecast, yoy_forecast, volume_forecast, driver_rows = _segment_projection(node, yaml_years)
            revenue_values = {**revenue_history, **revenue_forecast}
            volume_values = {**volume_history, **volume_forecast}
            segment_rows.append({"label": f"{segment_name} · 收入", "values": revenue_values, "format": "number", "bold": True})
            segment_rows.append({"label": f"{segment_name} · 同比", "values": _yoy_for_years(revenue_values, years), "format": "signed_percent", "muted": True})
            if _row_has_value(volume_values):
                segment_rows.append({"label": f"{segment_name} · 销量(万吨)", "values": volume_values, "format": "number", "muted": True})
            for driver, values in driver_rows.items():
                segment_rows.append(
                    {
                        "label": f"{segment_name} · {ASSUMPTION_LABELS.get(driver, driver)}",
                        "values": values,
                        "format": "decimal1" if driver in {"volume", "price", "revenue_yoy"} else "number",
                        "muted": True,
                        "assumption": True,
                    }
                )
    if segment_rows:
        row = _write_assumption_group(ws, row, title="主拆分 · 业务线", years=years, rows=segment_rows, base_period=base_period)

    used_paths: set[str] = set()
    for section_title, prefixes in ASSUMPTION_SECTION_DEFS:
        rows: list[dict[str, Any]] = []
        for path, payload in yaml1.items():
            if path in used_paths or not isinstance(payload, dict) or not any(path == prefix or path.startswith(prefix) for prefix in prefixes):
                continue
            values = payload.get("values")
            if not isinstance(values, list):
                continue
            row_values = {year: _historical_assumption_value(metrics, statements, path, year) for year in years if base_period and int(year) <= base_period}
            row_values.update({year: _num(value) for year, value in zip(yaml_years, values)})
            rows.append(
                {
                    "path": path,
                    "label": _assumption_label(path),
                    "values": row_values,
                    "format": "percent" if any(token in path for token in ("rate", "ratio", "gpm")) else "number",
                    "bold": "主动覆盖" in str(payload.get("src") or ""),
                    "assumption": True,
                }
            )
            used_paths.add(path)
        if section_title == "费用率":
            rows.sort(key=lambda item: EXPENSE_SECTION_ORDER.get(str(item.get("path")), 100))
        if rows:
            row = _write_assumption_group(ws, row, title=section_title, years=years, rows=rows, base_period=base_period)

    terminal = yaml1.get("terminal") if isinstance(yaml1.get("terminal"), dict) else {}
    fade = terminal.get("fade") if isinstance(terminal.get("fade"), dict) else {}
    terminal_rows = [
        {"label": "显式预测末年", "values": {"terminal": terminal.get("explicit_end")}, "format": "integer", "assumption": True},
        {"label": "Fade 到年份", "values": {"terminal": fade.get("to_year")}, "format": "integer", "assumption": True},
        {"label": "永续增长率", "values": {"terminal": terminal.get("perpetual_growth")}, "format": "percent", "assumption": True},
    ]
    term_start = row
    ws.merge_cells(start_row=term_start, start_column=1, end_row=term_start, end_column=len(years) + 1)
    _group_style(ws.cell(term_start, 1, "终值假设"))
    row = term_start + 1
    for item in terminal_rows:
        ws.cell(row, 1, item["label"])
        ws.cell(row, 2, item["values"].get("terminal"))
        _assumption_label_style(ws.cell(row, 1))
        _assumption_data_style(ws.cell(row, 2), number_format=_number_format(item["format"]))
        row += 1
    _auto_widths(ws, len(years) + 3)


def _read_csv_table(path: Path) -> tuple[list[str], dict[str, dict[str, Any]]]:
    if not path.exists():
        return [], {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)
        years = [str(int(float(row["period"]))) for row in rows if row.get("period")]
        values: dict[str, dict[str, Any]] = {}
        for field in reader.fieldnames or []:
            if field == "period":
                continue
            values[field] = {str(int(float(row["period"]))): _num(row.get(field)) for row in rows if row.get("period")}
        return years, values


def _statement_order(table_name: str, values: dict[str, dict[str, Any]]) -> list[str]:
    stmt = _registry.statement_meta_for_table(table_name)
    if stmt is None:
        return list(values)
    ordered = [field for field in stmt.field_order if field in values]
    ordered.extend(field for field in values if field not in ordered)
    return ordered


def _statement_label(table_name: str, field: str) -> str:
    stmt = _registry.statement_meta_for_table(table_name)
    if stmt and field in stmt.labels:
        return stmt.labels[field]
    return _registry.field_label(field, field)


def _statement_role(table_name: str, key: str, field: str) -> tuple[str, bool, bool]:
    stmt = _registry.statement_meta_for_table(table_name)
    if stmt is None:
        return "", False, False
    category = stmt.field_categories.get(field, "")
    is_total = field in stmt.total_fields
    is_key = field in (STATEMENT_KEY_ROWS.get(key) or set())
    is_technical = field.startswith("qa_") or category in {"combo", "derived", "sub_item"}
    return category, is_total or is_key, is_technical


def _statement_row_nonzero(row_values: dict[str, Any]) -> bool:
    return any(_num(value) is not None and abs(float(value)) > 1e-9 for value in row_values.values())


def _derived_statement_rows(anchor: str, metrics: dict[str, Any], years: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for metric, label, fmt in INCOME_DERIVED_ROWS.get(anchor, []):
        values = {year: _metric_value(metrics, year, metric) for year in years}
        if _row_has_value(values):
            rows.append({"field": metric, "label": label, "values": values, "format": fmt, "metric": True})
    return rows


def _visible_statement_years(company_dir: Path, metrics: dict[str, Any], years: list[str]) -> list[str]:
    base_period = int(float(metrics.get("base_period") or 0))
    yaml1 = _load_yaml1(company_dir)
    terminal = yaml1.get("terminal") if isinstance(yaml1.get("terminal"), dict) else {}
    explicit_end = _num(terminal.get("explicit_end"))
    history = [year for year in years if int(year) <= base_period][-5:]
    forecast = [year for year in years if int(year) > base_period]
    if explicit_end is not None:
        forecast = [year for year in forecast if int(year) <= int(explicit_end)]
    return history + forecast


def _fill_full_statement_sheet(
    workbook: Any,
    company_dir: Path,
    metrics: dict[str, Any],
    *,
    table_name: str,
    title: str,
    key: str,
    statement_title: str,
    unit: str,
) -> None:
    years, values = _read_csv_table(company_forecast_dir(company_dir) / table_name)
    if not years:
        return
    years = _visible_statement_years(company_dir, metrics, years)
    base_period = int(float(metrics.get("base_period") or 0))
    max_col = len(years) + 1
    ws = _fresh_sheet(workbook, title)
    _style_report_sheet(ws, title=title, subtitle=f"单位：{unit}", max_col=max_col)
    _write_header(ws, 3, ["科目", *[f"{year}E" if int(year) > base_period else year for year in years]])
    hidden = STATEMENT_HIDDEN_ROWS.get(key, set())
    row_index = 4
    for field in _statement_order(table_name, values):
        if field in hidden:
            continue
        row_values = values[field]
        if not _statement_row_nonzero(row_values):
            continue
        _, bold, is_technical = _statement_role(table_name, key, field)
        if is_technical:
            continue
        ws.cell(row_index, 1, _statement_label(table_name, field))
        for col_offset, year in enumerate(years, 2):
            cell = ws.cell(row_index, col_offset, row_values.get(year))
            _data_style(cell, forecast=int(year) > base_period)
        _label_style(ws.cell(row_index, 1), bold=bold)
        if bold:
            _emphasize_row(ws, row_index, max_col)
        row_index += 1
        if key == "is":
            for metric_row in _derived_statement_rows(field, metrics, years):
                ws.cell(row_index, 1, metric_row["label"])
                for col_offset, year in enumerate(years, 2):
                    cell = ws.cell(row_index, col_offset, metric_row["values"].get(year))
                    _data_style(cell, forecast=int(year) > base_period, number_format=_number_format(metric_row["format"]))
                _label_style(ws.cell(row_index, 1), muted=True)
                row_index += 1
    _auto_widths(ws, max_col)


def _fill_full_statement_sheets(workbook: Any, company_dir: Path, metrics: dict[str, Any]) -> None:
    for table_name, title, key, statement_title, unit in FULL_STATEMENT_SHEETS:
        _fill_full_statement_sheet(
            workbook,
            company_dir,
            metrics,
            table_name=table_name,
            title=title,
            key=key,
            statement_title=statement_title,
            unit=unit,
        )


def _quarter_periods(metrics: dict[str, Any]) -> list[str]:
    quarterly = metrics.get("quarterly")
    if not isinstance(quarterly, dict):
        return []
    periods = []
    for period in quarterly.get("periods", []):
        text = str(period)
        if len(text) == 6 and text[:4].isdigit() and text[4] == "Q" and text[5].isdigit():
            periods.append(text)
    return sorted(set(periods), key=lambda item: (int(item[:4]), int(item[-1])))


def _semiannual_periods(metrics: dict[str, Any]) -> list[dict[str, Any]]:
    quarter_set = set(_quarter_periods(metrics))
    if not quarter_set:
        return []
    base_period = int(float(metrics.get("base_period") or 0))
    years = sorted({int(period[:4]) for period in quarter_set})
    visible_years = years[-4:]
    periods: list[dict[str, Any]] = []
    for year in visible_years:
        h1_quarters = [f"{year}Q1", f"{year}Q2"]
        h2_quarters = [f"{year}Q3", f"{year}Q4"]
        if all(period in quarter_set for period in h1_quarters):
            forecast = year > base_period
            periods.append(
                {
                    "key": f"{year}H1",
                    "year": year,
                    "kind": "H1",
                    "quarters": h1_quarters,
                    "label": f"{year}H1E" if forecast else f"{year}H1",
                    "forecast": forecast,
                }
            )
        if all(period in quarter_set for period in h2_quarters) or _value(metrics, year, "revenue") is not None:
            forecast = year > base_period
            periods.append(
                {
                    "key": f"{year}H2",
                    "year": year,
                    "kind": "H2",
                    "quarters": h2_quarters,
                    "label": f"{year}H2E" if forecast else f"{year}H2",
                    "forecast": forecast,
                }
            )
    return periods


def _quarter_amount(metrics: dict[str, Any], period: str, field: str) -> float | None:
    value = _quarter_raw_value(metrics, period, field)
    if value is None:
        value = _quarter_value(metrics, period, field)
    return _num(value)


def _semiannual_amount(metrics: dict[str, Any], period: dict[str, Any], field: str) -> float | None:
    year = period.get("year")
    total = 0.0
    has_value = False
    for quarter in period.get("quarters", []):
        value = _quarter_amount(metrics, str(quarter), field)
        if value is None:
            continue
        total += value
        has_value = True
    if has_value:
        return total
    if period.get("kind") == "H2":
        annual = _num(_value(metrics, year, field))
        h1 = _semiannual_amount(metrics, {"year": year, "kind": "H1", "quarters": [f"{year}Q1", f"{year}Q2"]}, field)
        if annual is not None and h1 is not None:
            return annual - h1
    return None


def _semiannual_revenue(metrics: dict[str, Any], period: dict[str, Any]) -> float | None:
    return _semiannual_value(metrics, period, {}, "revenue")


def _semiannual_value(
    metrics: dict[str, Any],
    period: dict[str, Any],
    period_map: dict[str, dict[str, Any]],
    metric: str,
) -> float | None:
    year = int(period.get("year") or 0)
    kind = str(period.get("kind") or "")
    if metric == "gross_profit":
        revenue = _semiannual_value(metrics, period, period_map, "revenue")
        cost = _semiannual_value(metrics, period, period_map, "oper_cost")
        return None if revenue is None or cost is None else revenue - cost
    if metric in SEMIANNUAL_AMOUNT_FIELDS:
        return _semiannual_amount(metrics, period, metric)
    if metric == "revenue_yoy":
        current = _semiannual_value(metrics, period, period_map, "revenue")
        previous = _semiannual_value(metrics, period_map.get(f"{year - 1}{kind}", {}), period_map, "revenue")
        return _growth_rate(current, previous)
    if metric == "n_income_attr_p_yoy":
        current = _semiannual_value(metrics, period, period_map, "n_income_attr_p")
        previous = _semiannual_value(metrics, period_map.get(f"{year - 1}{kind}", {}), period_map, "n_income_attr_p")
        return _growth_rate(current, previous)
    denominator = _semiannual_value(metrics, period, period_map, "revenue")
    ratio_numerators = {
        "gross_margin": "gross_profit",
        "sell_exp_rate": "sell_exp",
        "admin_exp_rate": "admin_exp",
        "rd_exp_rate": "rd_exp",
        "fin_exp_rate": "fin_exp",
        "operate_margin": "operate_profit",
        "total_profit_margin": "total_profit",
        "n_income_margin": "n_income",
        "n_income_attr_p_margin": "n_income_attr_p",
    }
    if metric in ratio_numerators:
        return _safe_div(_semiannual_value(metrics, period, period_map, ratio_numerators[metric]), denominator)
    if metric == "effective_tax_rate":
        return _safe_div(_semiannual_value(metrics, period, period_map, "income_tax"), _semiannual_value(metrics, period, period_map, "total_profit"))
    return None


def _write_multilevel_period_header(ws: Any, periods: list[dict[str, Any]]) -> int:
    max_col = 1 + len(periods) * 2
    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
    _header_style(ws.cell(3, 1, "项目"))
    for index, period in enumerate(periods):
        start_col = 2 + index * 2
        end_col = start_col + 1
        ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
        _header_style(ws.cell(3, start_col, period["label"]))
        for col, label in ((start_col, "金额"), (end_col, "比率")):
            cell = ws.cell(4, col, label)
            _header_style(cell)
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 18
    return max_col


def _breakdown_period_key(row: dict[str, Any]) -> str | None:
    year_value = _num(row.get("year"))
    if year_value is None:
        return None
    year = int(year_value)

    explicit_period = str(row.get("period") or "").strip()
    if re.match(r"^\d{4}(A|H1)$", explicit_period):
        return explicit_period

    period_type = str(row.get("period_type") or "").strip().lower()
    if period_type == "annual":
        return f"{year}A"
    if period_type == "h1":
        return f"{year}H1"

    source = " ".join(str(row.get(field) or "") for field in ("source_file", "source_section", "source_table"))
    is_half = any(token in source for token in ("半年度", "半年报", "中报", "半年度报告"))
    return f"{year}H1" if is_half else f"{year}A"


def _pct_ratio(value: Any) -> float | None:
    number = _num(value)
    if number is None:
        return None
    return number / 100.0


def _is_half_report(path: Path) -> bool:
    name = path.name
    return any(token in name for token in ("半年度", "半年报", "中报"))


def _read_revenue_breakdown_rows(company_dir: Path, *, max_year: int | None = None) -> list[dict[str, Any]]:
    breakdown_dir = official_breakdowns_dir(company_dir)
    all_path = breakdown_dir / "business_revenue_breakdown_all.csv"
    annual_path = breakdown_dir / "business_revenue_breakdown.csv"
    h1_path = breakdown_dir / "business_revenue_breakdown_h1.csv"

    def read_csv(path: Path) -> list[dict[str, Any]]:
        if not path.exists():
            return []
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))

    def within_max_year(row: dict[str, Any]) -> bool:
        if max_year is None:
            return True
        row_year = _num(row.get("year"))
        return row_year is None or int(row_year) <= max_year

    rows: list[dict[str, Any]] = []
    if all_path.exists():
        return [row for row in read_csv(all_path) if within_max_year(row)]
    if annual_path.exists():
        rows.extend(read_csv(annual_path))
    if h1_path.exists():
        rows.extend(read_csv(h1_path))
        return [row for row in rows if within_max_year(row)]

    try:
        from dataclasses import asdict

        from src.business_breakdown_extractor import extract_report
    except Exception:
        return [row for row in rows if within_max_year(row)]

    reports_dir = quarterly_reports_dir(company_dir)
    if not reports_dir.exists():
        return rows
    reports = []
    for report in sorted(path for path in reports_dir.rglob("*.md") if _is_half_report(path)):
        report_year = _num(report.name[:4])
        if max_year is not None and report_year is not None and int(report_year) > max_year:
            continue
        reports.append(report)

    def extract_one(report: Path) -> list[dict[str, Any]]:
        try:
            return [asdict(row) for row in extract_report(report, company_dir=company_dir)]
        except Exception:
            return []

    for report_rows in parallel_map(extract_one, reports, max_workers=min(6, len(reports) or 1)):
        rows.extend(report_rows)
    return [row for row in rows if within_max_year(row)]


def _breakdown_score(row: dict[str, Any]) -> int:
    table = str(row.get("source_table") or "")
    if table == "revenue_composition":
        return 0
    if table == "major_business_profitability":
        return 1
    return 2


def _revenue_split_sections(
    company_dir: Path,
    metrics: dict[str, Any],
    periods: list[dict[str, Any]],
) -> dict[str, dict[str, dict[str, tuple[float | None, float | None]]]]:
    period_map = {period["key"]: period for period in periods}
    base_period = int(float(metrics.get("base_period") or 0))
    sections: dict[str, dict[str, dict[str, tuple[float | None, float | None]]]] = {
        "整体收入": {"营业收入": {}},
    }
    for key, period in period_map.items():
        revenue = _semiannual_revenue(metrics, period)
        sections["整体收入"]["营业收入"][key] = (revenue, 1.0 if revenue is not None else None)

    selected: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in _read_revenue_breakdown_rows(company_dir):
        row_year = _num(row.get("year"))
        period_key = _breakdown_period_key(row)
        if period_key is None:
            continue
        if base_period and row_year is not None and int(row_year) > base_period and period_key.endswith("A"):
            continue
        if period_key not in period_map and not period_key.endswith("A"):
            continue
        section = str(row.get("dimension_label") or row.get("dimension") or "").strip()
        item = str(row.get("item_name") or "").strip()
        if not section or not item:
            continue
        key = (section, item, period_key)
        current = selected.get(key)
        if current is None or _breakdown_score(row) < _breakdown_score(current):
            selected[key] = row
        elif current is not None and current.get("revenue_pct") in (None, "") and row.get("revenue_pct") not in (None, ""):
            current["revenue_pct"] = row.get("revenue_pct")

    annual_values: dict[tuple[str, str, int], tuple[float | None, float | None]] = {}
    for (section, item, period_key), row in selected.items():
        revenue_yuan = _num(row.get("revenue_yuan"))
        revenue = None if revenue_yuan is None else revenue_yuan / 1_000_000.0
        share = _pct_ratio(row.get("revenue_pct"))
        year = int(period_key[:4])
        if period_key.endswith("A"):
            annual_values[(section, item, year)] = (revenue, share)
            continue
        if period_key in period_map:
            total_revenue = _semiannual_revenue(metrics, period_map[period_key])
            if share is None and revenue is not None and total_revenue is not None and abs(total_revenue) > 1e-9:
                share = revenue / total_revenue
            sections.setdefault(section, {}).setdefault(item, {})[period_key] = (revenue, share)

    for (section, item, year), (annual_revenue, _) in annual_values.items():
        h1_key = f"{year}H1"
        h2_key = f"{year}H2"
        if h2_key not in period_map:
            continue
        h1_revenue = sections.get(section, {}).get(item, {}).get(h1_key, (None, None))[0]
        if annual_revenue is None or h1_revenue is None:
            continue
        h2_revenue = annual_revenue - h1_revenue
        h2_total = _semiannual_revenue(metrics, period_map[h2_key])
        h2_share = None if h2_total is None or abs(h2_total) < 1e-9 else h2_revenue / h2_total
        sections.setdefault(section, {}).setdefault(item, {})[h2_key] = (h2_revenue, h2_share)
    return sections


def _sort_revenue_items(rows: dict[str, dict[str, tuple[float | None, float | None]]], periods: list[dict[str, Any]]) -> list[str]:
    if list(rows) == ["营业收入"]:
        return ["营业收入"]
    latest_keys = [period["key"] for period in reversed(periods)]

    def score(item: str) -> tuple[float, str]:
        values = rows.get(item, {})
        latest = 0.0
        for key in latest_keys:
            value = values.get(key, (None, None))[0]
            if value is not None:
                latest = float(value)
                break
        return (-latest, item)

    return sorted(rows, key=score)


def _blank_data_style(cell: Any, *, forecast: bool) -> None:
    _data_style(cell, forecast=forecast, number_format='General')


def _split_label_style(cell: Any) -> None:
    _label_style(cell, muted=True)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)


def _write_semiannual_metric_row(
    ws: Any,
    row_index: int,
    *,
    metrics: dict[str, Any],
    periods: list[dict[str, Any]],
    period_map: dict[str, dict[str, Any]],
    metric: str,
    label: str,
    fmt: str,
    bold: bool,
    muted: bool,
    max_col: int,
) -> None:
    is_ratio = fmt in {"percent", "signed_percent", "multiple"}
    ws.cell(row_index, 1, label)
    _label_style(ws.cell(row_index, 1), bold=bold, muted=muted)
    for index, period in enumerate(periods):
        amount_col = 2 + index * 2
        ratio_col = amount_col + 1
        forecast = bool(period["forecast"])
        value = _semiannual_value(metrics, period, period_map, metric)
        if is_ratio:
            _blank_data_style(ws.cell(row_index, amount_col, None), forecast=forecast)
            _data_style(ws.cell(row_index, ratio_col, value), forecast=forecast, number_format=_number_format(fmt))
        else:
            _data_style(ws.cell(row_index, amount_col, value), forecast=forecast, number_format=_number_format(fmt))
            ratio_value = 1.0 if metric == "revenue" and value is not None else None
            _data_style(ws.cell(row_index, ratio_col, ratio_value), forecast=forecast, number_format=_number_format("percent"))
    if bold:
        _emphasize_row(ws, row_index, max_col)


def _write_embedded_revenue_split(
    ws: Any,
    row_index: int,
    *,
    sections: dict[str, dict[str, dict[str, tuple[float | None, float | None]]]],
    periods: list[dict[str, Any]],
    max_col: int,
) -> int:
    for section in REVENUE_SPLIT_SECTION_ORDER:
        if section == "整体收入":
            continue
        rows = sections.get(section)
        if not rows:
            continue
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=max_col)
        _subsection_style(ws.cell(row_index, 1, f"{section}拆分"))
        row_index += 1
        for item in _sort_revenue_items(rows, periods):
            values = rows.get(item, {})
            ws.cell(row_index, 1, item)
            _split_label_style(ws.cell(row_index, 1))
            for index, period in enumerate(periods):
                amount, share = values.get(period["key"], (None, None))
                amount_col = 2 + index * 2
                ratio_col = amount_col + 1
                forecast = bool(period["forecast"])
                _data_style(ws.cell(row_index, amount_col, amount), forecast=forecast)
                _data_style(ws.cell(row_index, ratio_col, share), forecast=forecast, number_format=_number_format("percent"))
            row_index += 1
    return row_index


def _fill_semiannual_is_sheet(workbook: Any, company_dir: Path, metrics: dict[str, Any]) -> None:
    periods = _semiannual_periods(metrics)
    if not periods:
        return
    period_map = {period["key"]: period for period in periods}
    split_sections = _revenue_split_sections(company_dir, metrics, periods)
    max_col = 1 + len(periods) * 2
    ws = _fresh_sheet(workbook, SEMIANNUAL_IS_SHEET)
    _style_report_sheet(
        ws,
        title=SEMIANNUAL_IS_SHEET,
        subtitle="单位：百万元；收入拆分来自半年报/年报，利润率按半年度与全年口径重算",
        max_col=max_col,
    )
    ws.freeze_panes = "B5"
    _write_multilevel_period_header(ws, periods)
    row_index = 5
    for section, rows in SEMIANNUAL_IS_SECTIONS:
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=max_col)
        _group_style(ws.cell(row_index, 1, section))
        row_index += 1
        for metric, label, fmt, bold, muted in rows:
            values = {period["key"]: _semiannual_value(metrics, period, period_map, metric) for period in periods}
            if not _row_has_value(values):
                continue
            _write_semiannual_metric_row(
                ws,
                row_index,
                metrics=metrics,
                periods=periods,
                period_map=period_map,
                metric=metric,
                label=label,
                fmt=fmt,
                bold=bold,
                muted=muted,
                max_col=max_col,
            )
            row_index += 1
            if section == "收入与毛利" and metric == "revenue":
                row_index = _write_embedded_revenue_split(ws, row_index, sections=split_sections, periods=periods, max_col=max_col)
    _auto_widths(ws, max_col)
    ws.column_dimensions["A"].width = 28
    for col in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10


def _fill_semiannual_sheets(workbook: Any, company_dir: Path, metrics: dict[str, Any]) -> None:
    if SEMIANNUAL_REVENUE_SPLIT_SHEET in workbook.sheetnames:
        workbook.remove(workbook[SEMIANNUAL_REVENUE_SPLIT_SHEET])
    _fill_semiannual_is_sheet(workbook, company_dir, metrics)


def _fill_quarterly_is_sheet(workbook: Any, company_dir: Path, metrics: dict[str, Any]) -> None:
    """Quarterly IS tracking sheet — mirrors frontend QuarterlyTable in three-statement style."""
    from .quarterly_tracker import compute_quarterly_view

    db_path = company_db_path(company_dir)
    ticker = metrics.get("ticker") or _ticker_from_db(db_path)
    try:
        view = compute_quarterly_view(db=db_path, ticker=ticker, company_dir=company_dir, year=None)
    except Exception:
        return  # non-blocking: skip sheet if quarterly view unavailable (e.g. no forecast_is.csv)

    periods = view["periods"]              # 12 "YYYYQq"
    rows = view["rows"]                    # amount + metric rows, in field_order
    annual = view["annual"]                # {field: value} (incl. metric fields via annual_out)
    quarter_states = view["quarter_states"]  # {"1":..,"2":..,"3":..,"4":..}
    selected_year = int(view["year"])

    if not periods or not rows:
        return

    annual_col = 1 + len(periods) + 1      # 科目 + 12 季 + 年度
    max_col = annual_col
    ws = _fresh_sheet(workbook, QUARTERLY_IS_SHEET)
    _style_report_sheet(ws, title=QUARTERLY_IS_SHEET, subtitle="单位：百万元", max_col=max_col)
    ws.freeze_panes = "B6"                 # override _style_report_sheet's B4 (header is 5 rows)

    # --- 3-row header ---
    ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=1)
    _header_style(ws.cell(3, 1, "科目"))
    for group_idx, year in enumerate([selected_year - 2, selected_year - 1, selected_year]):
        start = 2 + group_idx * 4
        end = start + 3
        ws.merge_cells(start_row=3, start_column=start, end_row=3, end_column=end)
        label = f"{year}(选定)" if year == selected_year else str(year)
        _header_style(ws.cell(3, start, label))
    for idx, period in enumerate(periods):
        _header_style(ws.cell(4, 2 + idx, f"{int(period[-1])}Q"))
    ws.merge_cells(start_row=3, start_column=annual_col, end_row=5, end_column=annual_col)
    _header_style(ws.cell(3, annual_col, "年度"))
    # state badge row (row 5, selected-year 4 cols only)
    for idx, period in enumerate(periods):
        if int(period[:4]) != selected_year:
            continue
        q = int(period[-1])
        state = quarter_states.get(str(q), "inherit")
        badge, color = QUARTERLY_STATE_BADGE.get(state, ("", MODEL_GREY))
        cell = ws.cell(5, 2 + idx, badge)
        cell.font = Font(name=MODEL_FONT, color=color, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor=MODEL_FORECAST_FILL)
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 16

    # --- data rows from row 6 ---
    row_index = 6
    for row in rows:
        role = row.get("role")
        fmt = row.get("format", "number")
        is_metric = role == "metric"
        is_key = role == "total" or bool(row.get("highlight"))
        _label_style(ws.cell(row_index, 1, row.get("label", "")), muted=is_metric, bold=is_key)
        for idx, period in enumerate(periods):
            col = 2 + idx
            forecast = int(period[:4]) == selected_year
            value = row.get("values", {}).get(period)
            _data_style(ws.cell(row_index, col, value), forecast=forecast, number_format=_number_format(fmt))
        annual_value = annual.get(row.get("field"))
        annual_cell = ws.cell(row_index, annual_col, annual_value)
        _data_style(annual_cell, forecast=False, number_format=_number_format(fmt))
        annual_cell.fill = PatternFill("solid", fgColor=MODEL_SUBTLE_FILL)
        if is_key:
            _emphasize_quarterly_row(ws, row_index, periods, annual_col, selected_year)
        row_index += 1

    _auto_widths(ws, max_col)
    for col in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11


def _value(metrics: dict[str, Any], period: int | str, metric: str) -> Any:
    if metric == "gross_sell_spread":
        gross_margin = _value(metrics, period, "gross_margin")
        sell_exp_rate = _value(metrics, period, "sell_exp_rate")
        if isinstance(gross_margin, (int, float)) and isinstance(sell_exp_rate, (int, float)):
            return gross_margin - sell_exp_rate
        return None
    annual = metrics.get("annual", {})
    row = annual.get(str(period), {}) if isinstance(annual, dict) else {}
    return row.get(metric) if isinstance(row, dict) else None


def _quarter_value(metrics: dict[str, Any], period: str, metric: str) -> Any:
    if metric == "gross_sell_spread":
        gross_margin = _quarter_value(metrics, period, "gross_margin")
        sell_exp_rate = _quarter_value(metrics, period, "sell_exp_rate")
        if isinstance(gross_margin, (int, float)) and isinstance(sell_exp_rate, (int, float)):
            return gross_margin - sell_exp_rate
        return None
    quarterly = metrics.get("quarterly")
    if not isinstance(quarterly, dict):
        return None
    by_period = quarterly.get("metrics_by_period", {})
    row = by_period.get(period, {}) if isinstance(by_period, dict) else {}
    return row.get(metric) if isinstance(row, dict) else None


def _year_label(year: int, base_period: int) -> int | str:
    return year if year <= base_period else f"{year}E"


def _format_quarter(period: str) -> str:
    if len(period) == 6 and period[:4].isdigit() and period[-2] == "Q":
        return f"{period[-1]}Q{period[:4]}"
    return period


def _set(ws: Any, cell: str, value: Any) -> None:
    ws[cell] = value


def _num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number


def _safe_div(numerator: Any, denominator: Any) -> float | None:
    num = _num(numerator)
    den = _num(denominator)
    if num is None or den is None or abs(den) < 1e-9:
        return None
    return num / den


def _read_period_records(path: Path) -> dict[str, dict[str, Any]]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = csv.DictReader(handle)
        out: dict[str, dict[str, Any]] = {}
        for row in rows:
            period = row.get("period")
            if period is None:
                continue
            year = str(int(float(period)))
            out[year] = {key: _num(value) for key, value in row.items() if key != "period"}
        return out


def _statement_records(company_dir: Path) -> dict[str, dict[str, dict[str, Any]]]:
    forecast_dir = company_forecast_dir(company_dir)
    return {
        "is": _read_period_records(forecast_dir / "full_is.csv"),
        "bs": _read_period_records(forecast_dir / "full_bs.csv"),
        "cf": _read_period_records(forecast_dir / "full_cf.csv"),
    }


def _quarterly_statement_records(company_dir: Path) -> dict[str, dict[str, Any]]:
    path = company_db_path(company_dir)
    if not path.exists():
        return {}
    connection = sqlite3.connect(path)
    connection.row_factory = sqlite3.Row
    try:
        rows = connection.execute("SELECT * FROM clean_quarterly ORDER BY period").fetchall()
    except sqlite3.Error:
        return {}
    finally:
        connection.close()
    out: dict[str, dict[str, Any]] = {}
    for row in rows:
        period = str(row["period"])
        out[period] = {key: _num(row[key]) if key != "period" else row[key] for key in row.keys() if key != "period"}
    return out


def _ticker_from_db(db_path: Path) -> str:
    """Read ticker from meta table (fallback when metrics omits it)."""
    if not db_path.exists():
        return ""
    conn = sqlite3.connect(db_path)
    try:
        row = conn.execute("SELECT value FROM meta WHERE key='ticker'").fetchone()
        return str(row[0]) if row else ""
    except sqlite3.Error:
        return ""
    finally:
        conn.close()


def _statement_value(statements: dict[str, dict[str, dict[str, Any]]], source: str, period: int | str, field: str) -> Any:
    return statements.get(source, {}).get(str(period), {}).get(field)


def _metric_value(metrics: dict[str, Any], period: int | str, metric: str) -> Any:
    return _value(metrics, period, metric)


def _previous_period(period: int | str) -> str:
    return str(int(period) - 1)


def _calc_value(metrics: dict[str, Any], statements: dict[str, dict[str, dict[str, Any]]], period: int | str, key: str) -> Any:
    p = str(period)
    prev = _previous_period(period)
    if key == "zero":
        return 0
    if key == "oper_cost_rate":
        return _safe_div(_metric_value(metrics, p, "oper_cost"), _metric_value(metrics, p, "revenue"))
    if key == "other_fin_exp":
        fin_exp = _metric_value(metrics, p, "fin_exp") or 0
        interest_exp = _statement_value(statements, "is", p, "fin_exp_int_exp") or 0
        interest_inc = _statement_value(statements, "is", p, "fin_exp_int_inc") or 0
        return fin_exp - (interest_exp - interest_inc)
    if key == "income_adjustments_total":
        return sum(
            _num(_statement_value(statements, "is", p, field)) or 0
            for field in ("oth_income", "invest_income", "fv_value_chg_gain", "asset_disp_income", "forex_gain")
        ) + (_metric_value(metrics, p, "non_operating_net") or 0)
    if key == "dividend_total":
        shares = _metric_value(metrics, p, "total_shares")
        dps = _metric_value(metrics, p, "dps")
        return None if shares is None or dps is None else shares * dps
    if key == "balance_check":
        assets = _metric_value(metrics, p, "total_assets") or 0
        liab = _metric_value(metrics, p, "total_liab") or 0
        equity = _metric_value(metrics, p, "total_equity") or 0
        return assets - liab - equity
    if key == "inventories_rate":
        return _safe_div(_statement_value(statements, "bs", p, "inventories"), _metric_value(metrics, p, "revenue"))
    if key == "receivables_total":
        return sum(
            _num(_statement_value(statements, "bs", p, field)) or 0
            for field in ("notes_receiv", "accounts_receiv", "receiv_financing", "prepayment", "oth_receiv")
        )
    if key == "receivables_rate":
        return _safe_div(_calc_value(metrics, statements, p, "receivables_total"), _metric_value(metrics, p, "revenue"))
    if key == "non_operating_current_liab":
        return sum(
            _num(_statement_value(statements, "bs", p, field)) or 0
            for field in ("st_borr", "st_fin_payable", "st_bonds_payable", "non_cur_liab_due_1y")
        )
    if key == "operating_wc_liabilities_rate":
        return _safe_div(_metric_value(metrics, p, "operating_wc_liabilities"), _metric_value(metrics, p, "revenue"))
    if key == "interest_bearing_ncl":
        return sum(_num(_statement_value(statements, "bs", p, field)) or 0 for field in ("lt_borr", "bond_payable", "lease_liab"))
    if key == "non_interest_ncl":
        return (_statement_value(statements, "bs", p, "total_ncl") or 0) - (_calc_value(metrics, statements, p, "interest_bearing_ncl") or 0)
    if key == "interest_expense_rate":
        return _safe_div(_statement_value(statements, "is", p, "fin_exp_int_exp"), _metric_value(metrics, p, "interest_bearing_debt"))
    if key == "interest_income_rate":
        return _safe_div(_statement_value(statements, "is", p, "fin_exp_int_inc"), _metric_value(metrics, p, "cash"))
    if key == "operating_assets_decrease":
        current = _metric_value(metrics, p, "operating_wc_assets")
        previous = _metric_value(metrics, prev, "operating_wc_assets")
        return None if current is None or previous is None else previous - current
    if key == "operating_liabilities_increase":
        current = _metric_value(metrics, p, "operating_wc_liabilities")
        previous = _metric_value(metrics, prev, "operating_wc_liabilities")
        return None if current is None or previous is None else current - previous
    if key == "defer_tax_assets_decrease":
        current = _statement_value(statements, "bs", p, "defer_tax_assets")
        previous = _statement_value(statements, "bs", prev, "defer_tax_assets")
        return None if current is None or previous is None else previous - current
    if key == "defer_tax_liab_increase":
        current = _statement_value(statements, "bs", p, "defer_tax_liab")
        previous = _statement_value(statements, "bs", prev, "defer_tax_liab")
        return None if current is None or previous is None else current - previous
    if key == "nopat":
        ebit = _metric_value(metrics, p, "ebit")
        tax_rate = _metric_value(metrics, p, "effective_tax_rate")
        return None if ebit is None or tax_rate is None else ebit * (1 - tax_rate)
    if key == "delta_nwc":
        current = _metric_value(metrics, p, "operating_nwc")
        previous = _metric_value(metrics, prev, "operating_nwc")
        return None if current is None or previous is None else current - previous
    if key == "fcff":
        nopat = _calc_value(metrics, statements, p, "nopat")
        da = _metric_value(metrics, p, "da")
        capex = _metric_value(metrics, p, "capex")
        delta_nwc = _calc_value(metrics, statements, p, "delta_nwc")
        if None in (nopat, da, capex, delta_nwc):
            return None
        return nopat + da - capex - delta_nwc
    if key == "cash_collection_quality":
        return _safe_div(_metric_value(metrics, p, "cfo"), _metric_value(metrics, p, "revenue"))
    if key == "cash_profit_quality":
        return _safe_div(_metric_value(metrics, p, "cfo"), _metric_value(metrics, p, "n_income_attr_p"))
    if key == "contract_liab_change_to_revenue":
        current = _statement_value(statements, "bs", p, "contract_liab")
        previous = _statement_value(statements, "bs", prev, "contract_liab")
        revenue = _metric_value(metrics, p, "revenue")
        return None if current is None or previous is None else _safe_div(current - previous, revenue)
    if key == "inventory_change_to_cogs":
        current = _statement_value(statements, "bs", p, "inventories")
        previous = _statement_value(statements, "bs", prev, "inventories")
        cogs = _metric_value(metrics, p, "oper_cost")
        return None if current is None or previous is None else _safe_div(current - previous, cogs)
    return None


def _model_bs_value(metrics: dict[str, Any], statements: dict[str, dict[str, dict[str, Any]]], period: int | str, row_index: int) -> Any:
    source = MODEL_BS_ROW_MAP.get(row_index)
    if source is None:
        return None
    source_type, field = source
    if source_type == "metric":
        return _metric_value(metrics, period, field)
    if source_type in {"is", "bs", "cf"}:
        return _statement_value(statements, source_type, period, field)
    if source_type == "calc":
        return _calc_value(metrics, statements, period, field)
    return None


def _quarter_raw_value(metrics: dict[str, Any], period: str, field: str) -> Any:
    quarterly = metrics.get("quarterly")
    if not isinstance(quarterly, dict):
        return None
    for row in quarterly.get("rows", []):
        if isinstance(row, dict) and row.get("field") == field:
            values = row.get("values", {})
            return values.get(period) if isinstance(values, dict) else None
    return None


MODEL_BS_STOCK_METRICS = {
    "cash",
    "interest_bearing_debt",
    "net_cash",
    "net_debt",
    "minority_int",
    "parent_equity",
    "total_equity",
    "total_assets",
    "total_liab",
    "total_shares",
    "operating_wc_assets",
    "operating_wc_liabilities",
    "operating_nwc",
    "invested_capital",
    "market_cap",
    "enterprise_value",
}


def _quarter_db_value(quarterly_records: dict[str, dict[str, Any]], period: str, field: str) -> Any:
    return quarterly_records.get(period, {}).get(field)


def _quarter_sum(
    metrics: dict[str, Any],
    quarterly_records: dict[str, dict[str, Any]],
    period: str,
    fields: tuple[str, ...],
) -> float | None:
    total = 0.0
    found = False
    for field in fields:
        value = _num(_quarter_metric_value(metrics, quarterly_records, period, field))
        if value is None:
            continue
        total += value
        found = True
    return total if found else None


def _quarter_base_lookup(
    metrics: dict[str, Any],
    quarterly_records: dict[str, dict[str, Any]],
    period: str,
    field: str,
) -> Any:
    """Non-recursive raw lookup for base IS fields (revenue/oper_cost/...).

    Stops at the three direct sources. Required because ``_quarter_metric_value``
    derives ratios from revenue/oper_cost — fetching those inputs via self-recursion
    infinite-loops when a quarter has no revenue in any source.
    """
    value = _quarter_value(metrics, period, field)
    if value is not None:
        return value
    value = _quarter_raw_value(metrics, period, field)
    if value is not None:
        return value
    return _quarter_db_value(quarterly_records, period, field)


def _quarter_metric_value(metrics: dict[str, Any], quarterly_records: dict[str, dict[str, Any]], period: str, field: str) -> Any:
    value = _quarter_base_lookup(metrics, quarterly_records, period, field)
    if value is not None:
        return value

    revenue = _num(_quarter_base_lookup(metrics, quarterly_records, period, "revenue"))
    oper_cost = _num(_quarter_base_lookup(metrics, quarterly_records, period, "oper_cost"))
    if field == "gross_profit":
        return None if revenue is None or oper_cost is None else revenue - oper_cost
    if field == "gross_margin":
        return _safe_div(_quarter_metric_value(metrics, quarterly_records, period, "gross_profit"), revenue)
    if field == "biz_tax_surchg_rate":
        return _safe_div(_quarter_metric_value(metrics, quarterly_records, period, "biz_tax_surchg"), revenue)
    if field == "sales_profit":
        gross_profit = _num(_quarter_metric_value(metrics, quarterly_records, period, "gross_profit"))
        tax = _num(_quarter_metric_value(metrics, quarterly_records, period, "biz_tax_surchg")) or 0.0
        return None if gross_profit is None else gross_profit - tax
    if field == "sales_profit_margin":
        return _safe_div(_quarter_metric_value(metrics, quarterly_records, period, "sales_profit"), revenue)
    if field in {"sell_exp_rate", "admin_exp_rate", "rd_exp_rate", "fin_exp_rate"}:
        return _safe_div(_quarter_metric_value(metrics, quarterly_records, period, field.removesuffix("_rate")), revenue)
    if field == "total_cogs":
        return _quarter_sum(metrics, quarterly_records, period, ("oper_cost", "biz_tax_surchg", "sell_exp", "admin_exp", "rd_exp", "fin_exp"))
    if field == "total_cogs_rate":
        return _safe_div(_quarter_metric_value(metrics, quarterly_records, period, "total_cogs"), revenue)
    if field == "sgna":
        return _quarter_sum(metrics, quarterly_records, period, ("sell_exp", "admin_exp"))
    if field == "sgna_rate":
        return _safe_div(_quarter_metric_value(metrics, quarterly_records, period, "sgna"), revenue)
    if field == "impairment":
        return _quarter_sum(metrics, quarterly_records, period, ("assets_impair_loss", "credit_impa_loss"))
    if field == "impairment_rate":
        return _safe_div(_quarter_metric_value(metrics, quarterly_records, period, "impairment"), revenue)
    if field == "ebit":
        operate_profit = _num(_quarter_metric_value(metrics, quarterly_records, period, "operate_profit"))
        fin_exp = _num(_quarter_metric_value(metrics, quarterly_records, period, "fin_exp")) or 0.0
        return None if operate_profit is None else operate_profit + fin_exp
    if field == "ebit_margin":
        return _safe_div(_quarter_metric_value(metrics, quarterly_records, period, "ebit"), revenue)
    if field == "da":
        return _quarter_sum(
            metrics,
            quarterly_records,
            period,
            ("depr_fa_coga_dpba", "amort_intang_assets", "lt_amort_deferred_exp", "use_right_asset_dep"),
        )
    if field == "ebitda":
        ebit = _num(_quarter_metric_value(metrics, quarterly_records, period, "ebit"))
        da = _num(_quarter_metric_value(metrics, quarterly_records, period, "da")) or 0.0
        return None if ebit is None else ebit + da
    if field == "ebitda_margin":
        return _safe_div(_quarter_metric_value(metrics, quarterly_records, period, "ebitda"), revenue)
    if field == "non_operating_net":
        income = _num(_quarter_metric_value(metrics, quarterly_records, period, "non_oper_income")) or 0.0
        expense = _num(_quarter_metric_value(metrics, quarterly_records, period, "non_oper_exp")) or 0.0
        return income - expense
    if field == "effective_tax_rate":
        return _safe_div(_quarter_metric_value(metrics, quarterly_records, period, "income_tax"), _quarter_metric_value(metrics, quarterly_records, period, "total_profit"))
    if field == "minority_gain":
        n_income = _num(_quarter_metric_value(metrics, quarterly_records, period, "n_income"))
        parent = _num(_quarter_metric_value(metrics, quarterly_records, period, "n_income_attr_p"))
        return None if n_income is None or parent is None else n_income - parent
    if field == "minority_gain_rate":
        return _safe_div(_quarter_metric_value(metrics, quarterly_records, period, "minority_gain"), revenue)
    if field in {"n_income_margin", "n_income_attr_p_margin"}:
        numerator = "n_income_attr_p"
        return _safe_div(_quarter_metric_value(metrics, quarterly_records, period, numerator), revenue)
    if field == "cash":
        return _quarter_db_value(quarterly_records, period, "money_cap")
    if field == "parent_equity":
        return _quarter_db_value(quarterly_records, period, "total_hldr_eqy_exc_min_int")
    if field == "total_equity":
        return _quarter_db_value(quarterly_records, period, "total_hldr_eqy_inc_min_int")
    if field == "total_shares":
        return _quarter_db_value(quarterly_records, period, "total_share")
    if field == "interest_bearing_debt":
        return _quarter_sum(metrics, quarterly_records, period, ("st_borr", "st_fin_payable", "st_bonds_payable", "non_cur_liab_due_1y", "lt_borr", "bond_payable", "lease_liab"))
    if field == "net_cash":
        cash = _num(_quarter_metric_value(metrics, quarterly_records, period, "cash"))
        debt = _num(_quarter_metric_value(metrics, quarterly_records, period, "interest_bearing_debt"))
        return None if cash is None or debt is None else cash - debt
    if field == "net_debt":
        net_cash = _num(_quarter_metric_value(metrics, quarterly_records, period, "net_cash"))
        return None if net_cash is None else -net_cash
    return None


def _model_period_quarters(period: str) -> list[str]:
    match = re.fullmatch(r"(\d{4})Q([1-4])", period)
    if match:
        return [period]
    match = re.fullmatch(r"(\d{4})H([12])", period)
    if not match:
        return []
    year = match.group(1)
    return [f"{year}Q1", f"{year}Q2"] if match.group(2) == "1" else [f"{year}Q3", f"{year}Q4"]


def _model_period_end_quarter(period: str) -> str | None:
    quarters = _model_period_quarters(period)
    return quarters[-1] if quarters else None


def _model_period_prior(period: str) -> str | None:
    match = re.fullmatch(r"(\d{4})([QH])([1-4])", period)
    if not match:
        return None
    return f"{int(match.group(1)) - 1}{match.group(2)}{match.group(3)}"


def _model_period_metric_value(metrics: dict[str, Any], quarterly_records: dict[str, dict[str, Any]], period: str, field: str) -> Any:
    quarters = _model_period_quarters(period)
    if not quarters:
        return None
    if len(quarters) == 1:
        return _quarter_metric_value(metrics, quarterly_records, period, field)
    end_quarter = quarters[-1]
    if field in MODEL_BS_STOCK_METRICS:
        return _quarter_metric_value(metrics, quarterly_records, end_quarter, field)
    if field in {"revenue_yoy", "oper_cost_yoy", "n_income_yoy", "n_income_attr_p_yoy"}:
        numerator = {
            "revenue_yoy": "revenue",
            "oper_cost_yoy": "oper_cost",
            "n_income_yoy": "n_income_attr_p",
            "n_income_attr_p_yoy": "n_income_attr_p",
        }[field]
        current = _model_period_metric_value(metrics, quarterly_records, period, numerator)
        prior = _model_period_prior(period)
        previous = _model_period_metric_value(metrics, quarterly_records, prior, numerator) if prior else None
        return None if current is None or previous is None or abs(float(previous)) < 1e-9 else float(current) / float(previous) - 1
    if field in {"gross_margin", "biz_tax_surchg_rate", "sales_profit_margin", "sell_exp_rate", "admin_exp_rate", "rd_exp_rate", "fin_exp_rate", "total_cogs_rate", "sgna_rate", "impairment_rate", "ebit_margin", "ebitda_margin", "effective_tax_rate", "minority_gain_rate", "n_income_margin", "n_income_attr_p_margin"}:
        if field == "gross_margin":
            return _safe_div(_model_period_metric_value(metrics, quarterly_records, period, "gross_profit"), _model_period_metric_value(metrics, quarterly_records, period, "revenue"))
        if field == "effective_tax_rate":
            return _safe_div(_model_period_metric_value(metrics, quarterly_records, period, "income_tax"), _model_period_metric_value(metrics, quarterly_records, period, "total_profit"))
        numerator = field.removesuffix("_rate").removesuffix("_margin")
        if field == "n_income_attr_p_margin":
            numerator = "n_income_attr_p"
        if field == "n_income_margin":
            numerator = "n_income_attr_p"
        if field == "sales_profit_margin":
            numerator = "sales_profit"
        if field == "total_cogs_rate":
            numerator = "total_cogs"
        if field == "minority_gain_rate":
            numerator = "minority_gain"
        return _safe_div(_model_period_metric_value(metrics, quarterly_records, period, numerator), _model_period_metric_value(metrics, quarterly_records, period, "revenue"))
    total = 0.0
    found = False
    for quarter in quarters:
        value = _num(_quarter_metric_value(metrics, quarterly_records, quarter, field))
        if value is None:
            continue
        total += value
        found = True
    return total if found else None


def _model_period_source_value(
    metrics: dict[str, Any],
    quarterly_records: dict[str, dict[str, Any]],
    period: str,
    source_type: str,
    field: str,
) -> Any:
    quarters = _model_period_quarters(period)
    if not quarters:
        return None
    if source_type == "bs":
        return _quarter_metric_value(metrics, quarterly_records, quarters[-1], field)
    if source_type in {"is", "cf"}:
        total = 0.0
        found = False
        for quarter in quarters:
            value = _num(_quarter_metric_value(metrics, quarterly_records, quarter, field))
            if value is None:
                continue
            total += value
            found = True
        return total if found else None
    return None


def _model_period_calc_value(metrics: dict[str, Any], quarterly_records: dict[str, dict[str, Any]], period: str, key: str) -> Any:
    if key == "zero":
        return 0
    if key == "oper_cost_rate":
        return _safe_div(_model_period_metric_value(metrics, quarterly_records, period, "oper_cost"), _model_period_metric_value(metrics, quarterly_records, period, "revenue"))
    if key == "other_fin_exp":
        fin_exp = _num(_model_period_metric_value(metrics, quarterly_records, period, "fin_exp")) or 0.0
        interest_exp = _num(_model_period_source_value(metrics, quarterly_records, period, "is", "fin_exp_int_exp")) or 0.0
        interest_inc = _num(_model_period_source_value(metrics, quarterly_records, period, "is", "fin_exp_int_inc")) or 0.0
        return fin_exp - (interest_exp - interest_inc)
    if key == "income_adjustments_total":
        total = sum(
            _num(_model_period_source_value(metrics, quarterly_records, period, "is", field)) or 0.0
            for field in ("oth_income", "invest_income", "fv_value_chg_gain", "asset_disp_income", "forex_gain")
        )
        return total + (_num(_model_period_metric_value(metrics, quarterly_records, period, "non_operating_net")) or 0.0)
    if key == "balance_check":
        assets = _num(_model_period_metric_value(metrics, quarterly_records, period, "total_assets")) or 0.0
        liab = _num(_model_period_metric_value(metrics, quarterly_records, period, "total_liab")) or 0.0
        equity = _num(_model_period_metric_value(metrics, quarterly_records, period, "total_equity")) or 0.0
        return assets - liab - equity
    if key == "inventories_rate":
        return _safe_div(_model_period_source_value(metrics, quarterly_records, period, "bs", "inventories"), _model_period_metric_value(metrics, quarterly_records, period, "revenue"))
    if key == "receivables_total":
        return sum(
            _num(_model_period_source_value(metrics, quarterly_records, period, "bs", field)) or 0.0
            for field in ("notes_receiv", "accounts_receiv", "receiv_financing", "prepayment", "oth_receiv")
        )
    if key == "receivables_rate":
        return _safe_div(_model_period_calc_value(metrics, quarterly_records, period, "receivables_total"), _model_period_metric_value(metrics, quarterly_records, period, "revenue"))
    if key in {"interest_expense_rate", "interest_income_rate", "operating_wc_liabilities_rate"}:
        numerator = {
            "interest_expense_rate": _model_period_source_value(metrics, quarterly_records, period, "is", "fin_exp_int_exp"),
            "interest_income_rate": _model_period_source_value(metrics, quarterly_records, period, "is", "fin_exp_int_inc"),
            "operating_wc_liabilities_rate": _model_period_metric_value(metrics, quarterly_records, period, "operating_wc_liabilities"),
        }[key]
        denominator = "cash" if key == "interest_income_rate" else "interest_bearing_debt" if key == "interest_expense_rate" else "revenue"
        return _safe_div(numerator, _model_period_metric_value(metrics, quarterly_records, period, denominator))
    return None


def _model_bs_period_value(metrics: dict[str, Any], quarterly_records: dict[str, dict[str, Any]], period: str, row_index: int) -> Any:
    source = MODEL_BS_ROW_MAP.get(row_index)
    if source is None:
        return None
    source_type, field = source
    if source_type == "metric":
        return _model_period_metric_value(metrics, quarterly_records, period, field)
    if source_type in {"is", "bs", "cf"}:
        return _model_period_source_value(metrics, quarterly_records, period, source_type, field)
    if source_type == "calc":
        return _model_period_calc_value(metrics, quarterly_records, period, field)
    return None


def _clear_cells(ws: Any, *, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _staticize_label_column(ws: Any) -> None:
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, 2).value
        if isinstance(value, str) and value.startswith("="):
            fallback = ws.cell(row, 1).value
            ws.cell(row, 2).value = str(fallback).strip() if fallback is not None else None


def _as_year(value: Any) -> int | None:
    if isinstance(value, (int, float)) and 1900 <= int(value) <= 2100:
        return int(value)
    if isinstance(value, str) and value.isdigit() and 1900 <= int(value) <= 2100:
        return int(value)
    return None


def _quarter_info(header_date: Any, original_label: Any) -> tuple[str, str] | None:
    if not isinstance(header_date, (datetime, date)):
        return None
    year = header_date.year
    month = header_date.month
    label_text = str(original_label or "")
    if month == 3:
        return f"{year}Q1", f"1Q{year}"
    if month == 6:
        if label_text.startswith("=IF"):
            return f"{year}H1", f"1H{year}"
        return f"{year}Q2", f"2Q{year}"
    if month == 9:
        return f"{year}Q3", f"3Q{year}"
    if month == 12:
        if label_text.startswith("=IF"):
            return f"{year}H2", f"2H{year}"
        return f"{year}Q4", f"4Q{year}"
    return None


def _period_year(period: str) -> int | None:
    match = re.match(r"^(\d{4})", period)
    return int(match.group(1)) if match else None


def _forecast_label(label: int | str, *, year: int | None, base_period: int) -> int | str:
    if year is None or year <= base_period:
        return label
    text = str(label)
    return text if text.endswith("E") else f"{text}E"


def _copy_cell_style(source: Any, target: Any) -> None:
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def _normalize_model_bs_forecast_styles(ws: Any, annual_columns: list[tuple[int, int]], base_period: int) -> None:
    forecast_columns = [col for col, year in annual_columns if year > base_period]
    if len(forecast_columns) < 2:
        return
    # The template's first forecast year can carry a one-off historical style.
    # Use the next forecast column as the house style and copy it across all forecast years.
    reference_col = forecast_columns[1]
    for col in forecast_columns:
        for row in range(2, ws.max_row + 1):
            source = ws.cell(row, reference_col)
            target = ws.cell(row, col)
            if isinstance(source, MergedCell) or isinstance(target, MergedCell):
                continue
            _copy_cell_style(source, target)


def _rating_report_year_specs() -> list[tuple[int, bool]]:
    config = app_config.rating_report_year_config()
    specs: list[tuple[int, bool]] = []
    seen: set[int] = set()
    for year in range(config["data_start_year"], config["data_end_year"] + 1):
        specs.append((year, False))
        seen.add(year)
    for year in range(config["forecast_start_year"], config["forecast_end_year"] + 1):
        if year in seen:
            continue
        specs.append((year, True))
        seen.add(year)
    return specs[: len(RATING_COLUMNS)]


def _rating_year_label(year: int, is_forecast: bool) -> int | str:
    return f"{year}E" if is_forecast else year


def _fill_summary(ws: Any, metrics: dict[str, Any], *, researcher_name: str = "") -> None:
    name = metrics.get("name") or metrics.get("ticker") or ""
    ticker = metrics.get("ticker") or ""
    base_period = int(float(metrics.get("base_period") or 0))
    _clear_cells(ws, min_row=7, max_row=100, min_col=4, max_col=ws.max_column)
    _clear_cells(ws, min_row=2, max_row=5, min_col=68, max_col=75)
    _set(ws, "B2", name)
    _set(ws, "BM2", ticker)
    _set(ws, "BE2", None)
    _set(ws, "B3", "研究员")
    _set(ws, "BM3", researcher_name or None)
    _set(ws, "B4", None)
    _set(ws, "BM4", None)

    for column, year in zip(SUMMARY_COLUMNS, SUMMARY_YEARS):
        label = _year_label(year, base_period)
        for header_row in (7, 9, 43, 58, 71):
            _set(ws, f"{column}{header_row}", label)
        for row_index, metric in SUMMARY_ROW_MAP.items():
            _set(ws, f"{column}{row_index}", _value(metrics, year, metric))


def _fill_model_bs(ws: Any, company_dir: Path, metrics: dict[str, Any], statements: dict[str, dict[str, dict[str, Any]]]) -> None:
    name = metrics.get("name") or metrics.get("ticker") or ""
    ticker = metrics.get("ticker") or ""
    base_period = int(float(metrics.get("base_period") or 0))
    available_years = {int(period) for period in metrics.get("periods", []) if str(period).isdigit()}
    quarterly_records = _quarterly_statement_records(company_dir)

    annual_columns: list[tuple[int, int]] = []
    quarter_columns: list[tuple[int, str, str]] = []
    for col in range(4, ws.max_column + 1):
        year = _as_year(ws.cell(2, col).value)
        if year is not None:
            annual_columns.append((col, year))
            continue
        info = _quarter_info(ws.cell(3, col).value, ws.cell(4, col).value)
        if info:
            period, label = info
            quarter_columns.append((col, period, label))

    _clear_cells(ws, min_row=2, max_row=ws.max_row, min_col=3, max_col=ws.max_column)
    _staticize_label_column(ws)
    _normalize_model_bs_forecast_styles(ws, annual_columns, base_period)
    _set(ws, "B1", "博时基金")
    _set(ws, "B3", ticker or name)
    _set(ws, "U1", date.today())

    for col, year in annual_columns:
        if year not in available_years:
            continue
        letter = get_column_letter(col)
        year_label = _forecast_label(year, year=year, base_period=base_period)
        _set(ws, f"{letter}2", year_label)
        _set(ws, f"{letter}3", datetime(year, 12, 31))
        _set(ws, f"{letter}4", year_label)
        for row_index in MODEL_BS_ROW_MAP:
            _set(ws, f"{letter}{row_index}", _model_bs_value(metrics, statements, year, row_index))

    for col, period, label in quarter_columns:
        letter = get_column_letter(col)
        year = _period_year(period)
        _set(ws, f"{letter}4", _forecast_label(label, year=year, base_period=base_period))
        for row_index in MODEL_BS_ROW_MAP:
            _set(ws, f"{letter}{row_index}", _model_bs_period_value(metrics, quarterly_records, period, row_index))


def _fill_rating(ws: Any, metrics: dict[str, Any]) -> None:
    year_specs = _rating_report_year_specs()
    _clear_cells(ws, min_row=1, max_row=11, min_col=2, max_col=9)
    for column, (year, is_forecast) in zip(RATING_COLUMNS, year_specs):
        _set(ws, f"{column}1", _rating_year_label(year, is_forecast))
        for row_index, metric in RATING_ROW_MAP.items():
            _set(ws, f"{column}{row_index}", _value(metrics, year, metric))


def _fill_comment_annual(ws: Any, metrics: dict[str, Any]) -> None:
    base_period = int(float(metrics.get("base_period") or 0))
    name = metrics.get("name") or metrics.get("ticker") or ""
    years = list(range(base_period - 4, base_period + 4))
    _clear_cells(ws, min_row=1, max_row=82, min_col=3, max_col=25)
    _set(ws, "B1", f"{name}-年度财务UE")
    for column, year in zip(COMMENT_ANNUAL_COLUMNS, years):
        _set(ws, f"{column}3", _year_label(year, base_period))
        for row_index, metric in COMMENT_ANNUAL_ROW_MAP.items():
            _set(ws, f"{column}{row_index}", _value(metrics, year, metric))


def _fill_comment_quarterly(ws: Any, metrics: dict[str, Any]) -> None:
    name = metrics.get("name") or metrics.get("ticker") or ""
    _set(ws, "B28", f"{name}-季度财务UE")
    quarterly = metrics.get("quarterly")
    periods = []
    if isinstance(quarterly, dict):
        periods = [str(period) for period in quarterly.get("periods", [])]
    periods = periods[-len(COMMENT_QUARTERLY_COLUMNS) :]
    for column, period in zip(COMMENT_QUARTERLY_COLUMNS, periods):
        _set(ws, f"{column}30", _format_quarter(period))
        for row_index, metric in COMMENT_QUARTERLY_ROW_MAP.items():
            _set(ws, f"{column}{row_index}", _quarter_value(metrics, period, metric))


def _fill_comment_rating_and_check(ws: Any, metrics: dict[str, Any]) -> None:
    name = metrics.get("name") or metrics.get("ticker") or ""
    base_period = int(float(metrics.get("base_period") or 0))
    _set(ws, "B59", f"{name}-评级模板")
    _set(ws, "B75", f"{name}-投资快速检查")
    _set(ws, "B80", None)
    _set(ws, "C80", name)

    rating_years = list(range(base_period - 3, base_period + 5))
    for column, year in zip(COMMENT_RATING_COLUMNS, rating_years):
        _set(ws, f"{column}62", _year_label(year, base_period))
    for offset, (rating_row, metric) in enumerate(RATING_ROW_MAP.items(), start=63):
        _set(ws, f"B{offset}", ws.parent.worksheets[3][f"A{rating_row}"].value)
        for column, year in zip(COMMENT_RATING_COLUMNS, rating_years):
            _set(ws, f"{column}{offset}", _value(metrics, year, metric))

    # Keep the quick-check block compact: current base year and first forecast year.
    forecast_year = base_period + 1
    for cell, year in (("D76", base_period), ("E76", base_period), ("F76", base_period), ("G76", base_period)):
        _set(ws, cell, year)
    for cell, year in (("H76", forecast_year), ("I76", forecast_year), ("J76", forecast_year), ("K76", forecast_year)):
        _set(ws, cell, year)
    _set(ws, "D80", (_value(metrics, base_period, "revenue") or 0) / 100.0 if _value(metrics, base_period, "revenue") is not None else None)
    _set(ws, "E80", _value(metrics, base_period, "revenue_yoy"))
    _set(ws, "F80", (_value(metrics, base_period, "n_income_attr_p") or 0) / 100.0 if _value(metrics, base_period, "n_income_attr_p") is not None else None)
    _set(ws, "G80", _value(metrics, base_period, "n_income_attr_p_yoy"))
    _set(ws, "H80", (_value(metrics, forecast_year, "revenue") or 0) / 100.0 if _value(metrics, forecast_year, "revenue") is not None else None)
    _set(ws, "I80", _value(metrics, forecast_year, "revenue_yoy"))
    _set(ws, "J80", (_value(metrics, forecast_year, "n_income_attr_p") or 0) / 100.0 if _value(metrics, forecast_year, "n_income_attr_p") is not None else None)
    _set(ws, "K80", _value(metrics, forecast_year, "n_income_attr_p_yoy"))
    market_cap = _value(metrics, base_period, "market_cap")
    _set(ws, "N80", market_cap / 100.0 if isinstance(market_cap, (int, float)) else None)
    _set(ws, "L80", _value(metrics, base_period, "pe"))
    _set(ws, "M80", _value(metrics, forecast_year, "pe"))
    base_profit = _value(metrics, base_period, "n_income_attr_p")
    forecast_profit = _value(metrics, forecast_year, "n_income_attr_p")
    base_yoy = _value(metrics, base_period, "n_income_attr_p_yoy")
    forecast_yoy = _value(metrics, forecast_year, "n_income_attr_p_yoy")
    base_pe = _value(metrics, base_period, "pe")
    forecast_pe = _value(metrics, forecast_year, "pe")
    if all(isinstance(value, (int, float)) for value in (base_profit, forecast_profit, base_yoy, forecast_yoy, base_pe, forecast_pe)):
        _set(
            ws,
            "D82",
            (
                f"{base_period}年利润 {base_profit / 100:.1f}亿，增速 {base_yoy:.0%}；"
                f"{forecast_year}年利润 {forecast_profit / 100:.1f}亿，增速 {forecast_yoy:.0%}；"
                f"对应估值 {base_pe:.0f}倍、{forecast_pe:.0f}倍。"
            ),
        )


def export_company_excel(
    company_dir: str | Path,
    *,
    metrics: dict[str, Any] | None = None,
    template_path: str | Path | None = None,
    output_path: str | Path | None = None,
) -> Path:
    """Create a value-filled Boshi presentation workbook without altering the template."""
    company_path = Path(company_dir)
    metrics = metrics or _load_metrics(company_path)
    template = Path(template_path) if template_path else DEFAULT_TEMPLATE_PATH
    if not template.exists():
        raise FileNotFoundError(f"Boshi Excel template not found: {template}")

    out_path = Path(output_path) if output_path else _default_output_path(company_path, metrics)
    explicit_output_path = output_path is not None
    out_path.parent.mkdir(parents=True, exist_ok=True)
    researcher_name = app_config.get_researcher_name()

    workbook = load_workbook(template, keep_vba=False, data_only=False, keep_links=False)
    workbook.defined_names.clear()
    statements = _statement_records(company_path)
    _fill_summary(_sheet(workbook, "Summary", 0), metrics, researcher_name=researcher_name)
    _fill_model_bs(_sheet(workbook, "Model-BS", 1), company_path, metrics, statements)
    _fill_rating(_sheet(workbook, "评级报告模板", 3), metrics)
    _remove_comment_sheets(workbook)
    _fill_core_assumptions_sheet(workbook, company_path, metrics, statements)
    _fill_full_statement_sheets(workbook, company_path, metrics)
    _fill_quarterly_is_sheet(workbook, company_path, metrics)
    _fill_semiannual_sheets(workbook, company_path, metrics)
    _apply_output_identity(workbook, researcher_name)

    with tempfile.TemporaryDirectory(prefix="boshi_excel_") as tmp_dir:
        temp_path = Path(tmp_dir) / out_path.name
        workbook.save(temp_path)
        _sanitize_workbook_package(temp_path)
        try:
            shutil.copy2(temp_path, out_path)
        except PermissionError:
            if explicit_output_path:
                raise
            out_path = _numbered_output_path(out_path)
            shutil.copy2(temp_path, out_path)
    return out_path
