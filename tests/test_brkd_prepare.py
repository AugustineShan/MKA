from __future__ import annotations

from pathlib import Path

import fitz
from openpyxl import Workbook

from src.brkd_prepare import prepare_brkd_materials
from src.company_paths import brkd_markdown_store_dir, brkd_material_dir


def _make_pdf(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((72, 72), text, fontsize=12, fontname="china-s")
    doc.save(path)
    doc.close()


def _make_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "收入"
    ws["B1"] = 100
    wb.save(path)


def test_prepare_brkd_materials_writes_markdown_store(tmp_path: Path):
    company = tmp_path / "companies" / "新乳业_002946"
    source_dir = brkd_material_dir(company)
    _make_pdf(source_dir / "研报.pdf", "hello research 业务")
    (source_dir / "纪要.txt").write_text("管理层纪要", encoding="utf-8")
    _make_workbook(source_dir / "表格.xlsx")

    manifest = prepare_brkd_materials(company, companies_dir=tmp_path / "companies")
    store = brkd_markdown_store_dir(company)

    assert manifest["source_count"] == 3
    assert store.is_dir()
    assert (store / "研报__pdf.md").exists()
    assert (store / "纪要__txt.md").exists()
    assert (store / "表格__xlsx.md").exists()
    assert (store / "brkd_prepare_manifest.json").exists()
    assert "hello research 业务" in (store / "研报__pdf.md").read_text(encoding="utf-8")
    assert "管理层纪要" in (store / "纪要__txt.md").read_text(encoding="utf-8")
    assert "Sheet: Summary" in (store / "表格__xlsx.md").read_text(encoding="utf-8")


def test_prepare_brkd_materials_idempotent_skips_unchanged(tmp_path: Path):
    company = tmp_path / "companies" / "新乳业_002946"
    source_dir = brkd_material_dir(company)
    (source_dir / "纪要.txt").parent.mkdir(parents=True, exist_ok=True)
    (source_dir / "纪要.txt").write_text("v1", encoding="utf-8")

    first = prepare_brkd_materials(company, companies_dir=tmp_path / "companies")
    second = prepare_brkd_materials(company, companies_dir=tmp_path / "companies")

    assert first["counts"]["converted"] == 1
    assert second["counts"]["skipped"] == 1


def test_prepare_brkd_materials_records_unsupported_binary_formats(tmp_path: Path):
    company = tmp_path / "companies" / "新乳业_002946"
    source_dir = brkd_material_dir(company)
    source_dir.mkdir(parents=True)
    (source_dir / "旧版.doc").write_bytes(b"legacy doc")

    manifest = prepare_brkd_materials(company, companies_dir=tmp_path / "companies")
    output = brkd_markdown_store_dir(company) / "旧版__doc.md"

    assert manifest["counts"]["unsupported"] == 1
    assert "暂不支持确定性转换" in output.read_text(encoding="utf-8")
