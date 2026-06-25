"""User-facing forecast pipeline tests."""

from __future__ import annotations

import json
import math
import sqlite3
import zipfile
from datetime import date
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from conftest import copy_fixture_company
from src import yaml1_cleaner
from src.assumption_staleness import StaleAssumptionError
from src.company_paths import forecast_dir, modelking_dir
from src.forecast import run_company_forecast


def _copy_new_hope_dairy(tmp_path: Path) -> Path:
    # Frozen snapshot (see tests/conftest.py) — deterministic forecast inputs.
    return copy_fixture_company(tmp_path)


def test_run_company_forecast_hides_intermediates_and_rebuilds_forecast(tmp_path, monkeypatch):
    company_dir = _copy_new_hope_dairy(tmp_path)
    monkeypatch.setattr(yaml1_cleaner, "COMPANIES_DIR", tmp_path / "companies")

    out_dir = forecast_dir(company_dir)
    internal_dir = modelking_dir(company_dir)
    stale_dir = out_dir
    stale_dir.mkdir(parents=True)
    (stale_dir / "stale.txt").write_text("old output", encoding="utf-8")

    run = run_company_forecast(ticker="002946.SZ")

    assert run.output_dir == out_dir
    assert not (company_dir / "forecast_params.yaml").exists()
    assert not (company_dir / "yaml1_clean_report.json").exists()
    assert not (company_dir / "yaml2_yearly.yaml").exists()
    assert not (stale_dir / "stale.txt").exists()

    assert (internal_dir / "forecast_params.yaml").exists()
    assert (internal_dir / "yaml1_clean_report.json").exists()
    assert (internal_dir / "forecast_build.json").exists()
    assert (out_dir / "forecast_is.csv").exists()
    assert (out_dir / "forecast_bs.csv").exists()
    assert (out_dir / "forecast_cf.csv").exists()
    assert (out_dir / "full_is.csv").exists()
    assert (out_dir / "full_bs.csv").exists()
    assert (out_dir / "full_cf.csv").exists()
    assert (out_dir / "dcf_summary.json").exists()
    assert (out_dir / "derived_metrics.json").exists()
    assert (out_dir / "derived_metrics_annual.csv").exists()
    assert (out_dir / "derived_metrics_quarterly.csv").exists()

    # Invariant assertions (not a brittle golden point value coupled to mutable
    # company data): the engine must produce a finite, positive, sanely-bounded
    # per-share value, and the projected balance sheet must balance every year.
    # The historical backtest hard-gate is asserted via the manifest below.
    summary = json.loads((out_dir / "dcf_summary.json").read_text(encoding="utf-8"))
    per_share = summary["per_share_value"]
    assert math.isfinite(per_share)
    assert 1.0 < per_share < 200.0

    bs = pd.read_csv(out_dir / "forecast_bs.csv")
    residual = (bs["total_assets"] - bs["total_liab"] - bs["total_hldr_eqy_inc_min_int"]).abs()
    assert (residual < 1.0).all(), f"projected balance sheet does not balance: max residual {residual.max()}"

    manifest = json.loads((out_dir / "run_manifest.json").read_text(encoding="utf-8"))
    assert manifest["contract"] == "Agent/defaults.yaml + Agent/yaml1*.yaml -> Agent/forecast/"
    assert manifest["yaml2_defaults_path"].endswith("defaults.yaml")
    # Naming-scheme-agnostic: the manifest must record a yaml1 file, not a specific stem.
    assert Path(manifest["yaml1_path"]).name.startswith("yaml1")
    assert manifest["yaml1_path"].endswith(".yaml")
    assert ".modelking" in manifest["internal_forecast_params_path"]
    assert manifest["backtest_status"] == "passed"
    assert manifest["derived_metrics_path"].endswith("derived_metrics.json")
    assert manifest["company_excel_export_status"] == "written"
    output_path = Path(manifest["company_excel_output_path"])
    assert output_path.parent == company_dir
    assert output_path.suffix == ".xlsx"
    assert output_path.name.endswith(f"Model-{date.today():%y%m%d}.xlsx")
    assert output_path.exists()

    workbook = load_workbook(output_path, data_only=False)
    assert "Summary" in workbook.sheetnames
    assert len(workbook.worksheets) == 8
    assert not any(title in {"点评模板", "\ub4d0\ud300\uce5c\uacbc"} for title in workbook.sheetnames)
    assert "半年度收入拆分" not in workbook.sheetnames
    assert {"核心假设", "完整利润表", "完整资产负债表", "完整现金流量表", "半年度利润表"}.issubset(
        set(workbook.sheetnames)
    )
    formula_cells = [
        cell.coordinate
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    assert formula_cells == []
    model_bs = workbook["Model-BS"]
    assert model_bs["BL5"].value is not None
    assert model_bs["BL345"].value is not None
    model_bs_headers = {
        str(model_bs.cell(4, column).value): column
        for column in range(1, model_bs.max_column + 1)
        if model_bs.cell(4, column).value is not None
    }
    for header in ("1Q2021", "2Q2021", "1H2021", "3Q2021", "4Q2021", "2H2021", "2025E"):
        assert header in model_bs_headers
    assert model_bs["BL2"].fill.fgColor.rgb == model_bs["BM2"].fill.fgColor.rgb
    assert model_bs["BL4"].fill.fgColor.rgb == model_bs["BM4"].fill.fgColor.rgb
    assert model_bs["BL2"].border.bottom.style == model_bs["BM2"].border.bottom.style
    q1_2021 = model_bs.cell(5, model_bs_headers["1Q2021"]).value
    q2_2021 = model_bs.cell(5, model_bs_headers["2Q2021"]).value
    h1_2021 = model_bs.cell(5, model_bs_headers["1H2021"]).value
    h2_2021 = model_bs.cell(5, model_bs_headers["2H2021"]).value
    assert isinstance(q1_2021, (float, int))
    assert isinstance(q2_2021, (float, int))
    assert isinstance(h1_2021, (float, int))
    assert isinstance(h2_2021, (float, int))
    assert h1_2021 == pytest.approx(q1_2021 + q2_2021)
    assert workbook["Summary"]["BP2"].value is None
    assert workbook["Summary"]["BR2"].value is None
    assert workbook["Summary"]["BU3"].value is None
    assert workbook["Summary"]["BV5"].value is None
    rating_sheet = workbook.worksheets[2]
    assert [rating_sheet.cell(1, column).value for column in range(2, 10)] == [
        2023,
        2024,
        2025,
        "2026E",
        "2027E",
        "2028E",
        None,
        None,
    ]
    assert rating_sheet["A9"].value == "PE"
    assert isinstance(rating_sheet["F9"].value, (float, int))
    assert workbook["核心假设"]["A5"].value == "营业收入"
    assert workbook["核心假设"]["B3"].value != "字段"
    core_sheet = workbook["核心假设"]
    assumption_cells = [
        core_sheet.cell(row, column)
        for row in range(1, core_sheet.max_row + 1)
        if core_sheet.cell(row, 1).value in {"整体毛利率", "销售费用率", "永续增长率"}
        for column in range(1, core_sheet.max_column + 1)
    ]
    assert any(cell.fill.fgColor.rgb == "00FFF2CC" for cell in assumption_cells)
    assert workbook["完整利润表"]["B3"].value != "字段"
    assert workbook["完整利润表"]["C3"].value != "分类"
    assert isinstance(workbook["完整利润表"]["B4"].value, (float, int))
    assert isinstance(workbook["完整资产负债表"]["B4"].value, (float, int))
    assert isinstance(workbook["完整现金流量表"]["B4"].value, (float, int))
    assert workbook["半年度利润表"]["A1"].value == "半年度利润表"
    assert "H1" in str(workbook["半年度利润表"]["B3"].value)
    assert "H2" in str(workbook["半年度利润表"]["D3"].value)
    assert "A" not in str(workbook["半年度利润表"]["D3"].value)
    assert workbook["半年度利润表"]["B4"].value == "金额"
    assert workbook["半年度利润表"]["C4"].value == "比率"
    assert isinstance(workbook["半年度利润表"]["B6"].value, (float, int))
    assert any(workbook["半年度利润表"].cell(row, 1).value == "收入与毛利" for row in range(1, workbook["半年度利润表"].max_row + 1))
    with zipfile.ZipFile(output_path) as archive:
        names = archive.namelist()
        assert not any(name.startswith("xl/externalLinks/") for name in names)
        assert not any("vbaProject" in name for name in names)
        rels = archive.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        workbook_xml = archive.read("xl/workbook.xml").decode("utf-8")
        content_types = archive.read("[Content_Types].xml").decode("utf-8")
        assert "externalLink" not in rels
        assert "externalReferences" not in workbook_xml
        assert "definedNames" not in workbook_xml
        assert "externalLink" not in content_types
        assert "vbaProject" not in content_types


def test_run_company_forecast_rejects_noncanonical_output_dir(tmp_path, monkeypatch):
    company_dir = _copy_new_hope_dairy(tmp_path)
    monkeypatch.setattr(yaml1_cleaner, "COMPANIES_DIR", tmp_path / "companies")

    with pytest.raises(Exception, match="must be named forecast"):
        run_company_forecast(ticker="002946.SZ", output_dir=company_dir / "Agent" / "forecast_yaml1")


def test_run_company_forecast_requires_annual_update_when_actual_overlaps_horizon(tmp_path, monkeypatch):
    company_dir = _copy_new_hope_dairy(tmp_path)
    monkeypatch.setattr(yaml1_cleaner, "COMPANIES_DIR", tmp_path / "companies")

    out_dir = forecast_dir(company_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    sentinel = out_dir / "stale.txt"
    sentinel.write_text("old forecast should survive blocked run", encoding="utf-8")

    db_path = company_dir / "Agent" / "data.db"
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute("select * from clean_annual where period = ?", ("2024",)).fetchone()
        assert row is not None
        columns = [item[1] for item in conn.execute("pragma table_info(clean_annual)").fetchall()]
        values = [row[column] for column in columns]
        values[columns.index("period")] = "2025"
        quoted_columns = ", ".join(f'"{column}"' for column in columns)
        placeholders = ", ".join("?" for _ in columns)
        conn.execute(
            f"insert or replace into clean_annual ({quoted_columns}) values ({placeholders})",
            values,
        )

    with pytest.raises(StaleAssumptionError, match="/annual-update") as exc_info:
        run_company_forecast(ticker="002946.SZ")

    status = exc_info.value.status
    assert status.data_end == 2025
    assert status.forecast_start == 2025
    assert status.defaults_base_year == 2024
    assert sentinel.exists()
