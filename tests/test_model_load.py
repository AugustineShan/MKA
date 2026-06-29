from __future__ import annotations

import sqlite3
from datetime import datetime
from pathlib import Path

import pytest
import yaml
from openpyxl import Workbook

from conftest import copy_fixture_company
from src.company_paths import annual_reports_dir, ka_reference_dir, load_model_dir
from src.model_load import (
    ModelLoadError,
    inspect_workbook_boundary,
    latest_active_model,
    prepare_load,
)


def _write_workbook(path: Path, labels: list[object], values: list[object] | None = None) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    raw = wb.create_sheet("Raw-IS")
    raw["A1"] = 100
    for idx, label in enumerate(labels, start=2):
        ws.cell(1, idx).value = label
    if values:
        for idx, value in enumerate(values, start=2):
            ws.cell(2, idx).value = value
    wb.save(path)
    return path


def _table_exists(conn: sqlite3.Connection, table: str) -> bool:
    row = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
        (table,),
    ).fetchone()
    return row is not None


def _copy_period_row(db_path: Path, table: str, source_period: str, dest_period: str) -> None:
    with sqlite3.connect(db_path) as conn:
        if not _table_exists(conn, table):
            return
        conn.row_factory = sqlite3.Row
        row = conn.execute(f'SELECT * FROM "{table}" WHERE period = ?', (source_period,)).fetchone()
        if row is None:
            row = conn.execute(f'SELECT * FROM "{table}" ORDER BY period DESC LIMIT 1').fetchone()
        if row is None:
            return
        columns = [item[1] for item in conn.execute(f'PRAGMA table_info("{table}")').fetchall()]
        values = [row[column] for column in columns]
        values[columns.index("period")] = dest_period
        quoted_columns = ", ".join(f'"{column}"' for column in columns)
        placeholders = ", ".join("?" for _ in columns)
        conn.execute(
            f'INSERT OR REPLACE INTO "{table}" ({quoted_columns}) VALUES ({placeholders})',
            values,
        )


def _periods(db_path: Path, table: str) -> set[str]:
    with sqlite3.connect(db_path) as conn:
        if not _table_exists(conn, table):
            return set()
        return {str(row[0]) for row in conn.execute(f'SELECT period FROM "{table}"').fetchall()}


def test_inspect_workbook_boundary_from_forecast_labels(tmp_path: Path):
    model = _write_workbook(tmp_path / "model20250527.xlsx", [2022, 2023, 2024, "2025E", "2026E"])

    boundary = inspect_workbook_boundary(model)

    assert boundary.model_asof_date == "2025-05-27"
    assert boundary.history_end_year == 2024
    assert boundary.forecast_start_year == 2025
    assert boundary.forecast_years == (2025, 2026)


def test_inspect_workbook_boundary_from_formula_switch(tmp_path: Path):
    model = _write_workbook(
        tmp_path / "model.xlsx",
        [2022, 2023, 2024, 2025],
        ["='Raw-IS'!A1", "='Raw-IS'!A1", "='Raw-IS'!A1", 123.0],
    )

    boundary = inspect_workbook_boundary(model)

    assert boundary.history_end_year == 2024
    assert boundary.forecast_start_year == 2025
    assert "switches to literals" in (boundary.source.formula_switch_detail or "")


def test_prepare_stops_when_label_and_formula_boundary_conflict(tmp_path: Path):
    company_dir = tmp_path / "companies" / "测试公司_000001"
    model = _write_workbook(
        load_model_dir(company_dir) / "model.xlsx",
        [2023, 2024, "2025E", "2026E"],
        ["='Raw-IS'!A1", "='Raw-IS'!A1", "='Raw-IS'!A1", 123.0],
    )

    with pytest.raises(ModelLoadError, match="explicit forecast label starts"):
        prepare_load(company_dir, model_path=model, load_id="conflict")


def test_prepare_load_builds_cutoff_db_defaults_and_forbidden_report(tmp_path: Path):
    company_dir = copy_fixture_company(tmp_path)
    model = _write_workbook(load_model_dir(company_dir) / "model20250527.xlsx", [2022, 2023, 2024, "2025E"])
    report_2025 = annual_reports_dir(company_dir) / "2025年度报告.md"
    report_2025.parent.mkdir(parents=True, exist_ok=True)
    report_2025.write_text("future report", encoding="utf-8")

    source_db = company_dir / "Agent" / "data.db"
    _copy_period_row(source_db, "clean_annual", "2024", "2025")
    _copy_period_row(source_db, "clean_quarterly", "2024Q4", "2025Q1")

    result = prepare_load(company_dir, model_path=model, load_id="case", overwrite=True)
    load_dir = Path(result["load_dir"])
    cutoff_db = load_dir / "data_cutoff.db"

    assert "2025" in _periods(source_db, "clean_annual")
    assert "2025" not in _periods(cutoff_db, "clean_annual")
    assert not any(period.startswith("2025") for period in _periods(cutoff_db, "clean_quarterly"))

    defaults = yaml.safe_load((load_dir / "defaults.yaml").read_text(encoding="utf-8"))
    assert str(defaults["base_period"]) == "2024"
    expected_name = f"核心假设参考load_{datetime.now().strftime('%Y%m%d')}.md"
    assert (load_dir / expected_name).exists()
    assert not (company_dir / expected_name).exists()
    assert ka_reference_dir(company_dir).is_dir()
    assert not (load_dir / "model20250527_核心假设.md").exists()
    assert result["core_assumption_path"] == str(load_dir / expected_name)
    assert result["core_assumption_scaffold_path"] == str(load_dir / expected_name)
    assert result["root_core_assumption_path"] == str(ka_reference_dir(company_dir) / expected_name)
    forbidden = (load_dir / "forbidden_materials.md").read_text(encoding="utf-8")
    assert "2025年度报告.md" in forbidden
    assert result["removed_rows"]["clean_annual"] >= 1


def test_latest_active_model_ignores_webclaude_packaged_copy(tmp_path: Path):
    company_dir = tmp_path / "companies" / "测试公司_000001"
    active = _write_workbook(load_model_dir(company_dir) / "active.xlsx", [2023, 2024, "2025E"])
    packaged = _write_workbook(company_dir / "WEBCLAUDE" / "核心假设部分" / "02_活跃素材_packaged.xlsx", [2023, 2024, "2025E"])
    packaged.touch()

    assert latest_active_model(company_dir) == active


def test_latest_active_model_requires_single_load_material(tmp_path: Path):
    company_dir = tmp_path / "companies" / "测试公司_000001"
    _write_workbook(load_model_dir(company_dir) / "a.xlsx", [2023, 2024, "2025E"])
    _write_workbook(load_model_dir(company_dir) / "b.xlsx", [2023, 2024, "2025E"])

    with pytest.raises(ModelLoadError, match="exactly one Excel"):
        latest_active_model(company_dir)
